from __future__ import annotations

import math
import os
import re
import sys
import threading
import warnings
warnings.filterwarnings("ignore", message="'cgi' is deprecated.*", category=DeprecationWarning)
import cgi
import csv
import io
import json
import shutil
import subprocess
import tempfile
import webbrowser
from html import escape
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote, unquote, urlencode, urlparse

import pdfplumber
from openpyxl import load_workbook


IS_FROZEN = bool(getattr(sys, "frozen", False))
EXEC_DIR = Path(sys.executable).resolve().parent if IS_FROZEN else Path(__file__).resolve().parents[1]
BUNDLE_ROOT = Path(getattr(sys, "_MEIPASS", EXEC_DIR)).resolve() if IS_FROZEN else EXEC_DIR
ROOT = BUNDLE_ROOT
APP_DIR = BUNDLE_ROOT / "genoma_app" if IS_FROZEN else Path(__file__).resolve().parent
if IS_FROZEN and not APP_DIR.exists():
    APP_DIR = EXEC_DIR / "genoma_app"
if IS_FROZEN and sys.platform == "darwin":
    DEFAULT_DATA_DIR = Path.home() / "Library" / "Application Support" / "GenomaPDFStudio"
else:
    DEFAULT_DATA_DIR = EXEC_DIR / "data" if IS_FROZEN else APP_DIR
DATA_DIR = Path(os.environ.get("GENOMA_DATA_DIR", DEFAULT_DATA_DIR)).resolve()
DEFAULT_XLSX = Path(r"C:\Users\keytb\Downloads\DATA SOFIA GUANCHEZ (1).xlsx")
DEFAULT_PDF = Path(r"C:\Users\keytb\Downloads\MICRO EMILIANA MONTENEGRO.pdf")
RESOURCES = BUNDLE_ROOT / "GENOMA_NINOS3_extracted" / "StaticResources" / "RegisteredResources"
if IS_FROZEN and not RESOURCES.exists():
    RESOURCES = EXEC_DIR / "GENOMA_NINOS3_extracted" / "StaticResources" / "RegisteredResources"
UPLOADS = DATA_DIR / "uploads"
MANUAL_SOURCE = "__manual__"
MANUAL_CATALOG = DATA_DIR / "catalog_manual.tsv"
CONSIDERATIONS_FILE = DATA_DIR / "consideraciones.txt"
YEAST_PROFILE_FILE = DATA_DIR / "micobioma_manual.tsv"
SESSION_STATE_FILE = DATA_DIR / "session_state.json"
CLINICAL_SETTINGS_FILE = DATA_DIR / "clinical_settings.json"
PATHOGEN_SETTINGS_FILE = DATA_DIR / "pathogen_settings.json"
TABLE_HEADER_SETTINGS_FILE = DATA_DIR / "table_header_settings.json"
CHROME_CANDIDATES = [
    Path("/usr/bin/chromium"),
    Path("/usr/bin/chromium-browser"),
    Path("/usr/bin/google-chrome"),
    Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
    Path("/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge"),
    Path("/Applications/Chromium.app/Contents/MacOS/Chromium"),
    Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
    Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
    Path(r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"),
    Path(r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"),
]
if os.environ.get("CHROME_PATH"):
    CHROME_CANDIDATES.insert(0, Path(os.environ["CHROME_PATH"]))


@dataclass
class ExtractedReport:
    patient: dict[str, str]
    observation: str
    page1_rows: list[dict[str, str]]
    page2_rows: list[dict[str, Any]]
    catalog: dict[str, list[dict[str, Any]]]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def norm(value: str) -> str:
    value = clean_text(value).lower()
    value = value.replace("“", '"').replace("”", '"')
    value = re.sub(r"[^a-z0-9%/().:+\- ]+", "", value)
    return re.sub(r"\s+", " ", value).strip()


def load_catalog(path: Path) -> dict[str, list[dict[str, Any]]]:
    if str(path) == MANUAL_SOURCE:
        return load_manual_catalog()
    wb = load_workbook(path, data_only=True, read_only=True)
    out: dict[str, list[dict[str, Any]]] = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            out[ws.title] = []
            continue
        headers = [clean_text(h) for h in rows[0]]
        records: list[dict[str, Any]] = []
        for row in rows[1:]:
            record = {headers[i] or f"Column{i+1}": row[i] for i in range(min(len(headers), len(row)))}
            if any(v is not None and clean_text(v) for v in record.values()):
                records.append(record)
        out[ws.title] = records
    return out


def seed_data_file(target: Path, source: Path) -> None:
    if target.exists() or not source.exists():
        return
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)


def ensure_data_dir() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    UPLOADS.mkdir(parents=True, exist_ok=True)
    seed_data_file(MANUAL_CATALOG, APP_DIR / "catalog_manual.tsv")
    seed_data_file(CONSIDERATIONS_FILE, APP_DIR / "consideraciones.txt")
    seed_data_file(YEAST_PROFILE_FILE, APP_DIR / "micobioma_manual.tsv")
    load_clinical_settings()
    load_pathogen_settings()
    load_table_header_settings()


def load_session_state() -> dict[str, Any]:
    if not SESSION_STATE_FILE.exists():
        return {}
    try:
        data = json.loads(SESSION_STATE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


def save_session_state(
    *,
    pdf: Path | None = None,
    xlsx: Path | None = None,
    patient_overrides: dict[str, str] | None = None,
) -> None:
    state = load_session_state()
    if pdf is not None and str(pdf):
        state["pdf"] = str(pdf)
    if xlsx is not None and str(xlsx):
        state["xlsx"] = str(xlsx)
    if patient_overrides:
        saved_patient = state.get("patient_overrides", {})
        if not isinstance(saved_patient, dict):
            saved_patient = {}
        saved_patient.update({key: value for key, value in patient_overrides.items() if clean_text(value)})
        state["patient_overrides"] = saved_patient
    SESSION_STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    SESSION_STATE_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


DEFAULT_CLINICAL_SETTINGS = {
    "enterotype_number": "1",
    "enterotype_name": "Proteolítico",
    "enterotype_text": "El paciente es del grupo 1. Proteolítico.",
    "stool_macro": "• pH: Alcalina.\n• Color: Marrón\n• Consistencia: Duras\n• Restos de alimentos presente.",
    "stool_micro": "- Parásitos: No se observaron formas parasitarias.\n- Blastosporas (levaduras): No observadas.\n- Pseudohifas: No observadas.",
}


def load_clinical_settings() -> dict[str, str]:
    if CLINICAL_SETTINGS_FILE.exists():
        try:
            data = json.loads(CLINICAL_SETTINGS_FILE.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                merged = dict(DEFAULT_CLINICAL_SETTINGS)
                merged.update({key: clean_text(value) for key, value in data.items()})
                return merged
        except Exception:
            pass
    CLINICAL_SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
    CLINICAL_SETTINGS_FILE.write_text(json.dumps(DEFAULT_CLINICAL_SETTINGS, ensure_ascii=False, indent=2), encoding="utf-8")
    return dict(DEFAULT_CLINICAL_SETTINGS)


def save_clinical_settings(settings: dict[str, str]) -> None:
    current = load_clinical_settings()
    current.update({key: clean_text(value) for key, value in settings.items()})
    number = current.get("enterotype_number", "1") or "1"
    name = current.get("enterotype_name", "Proteolítico") or "Proteolítico"
    if not current.get("enterotype_text"):
        current["enterotype_text"] = f"El paciente es del grupo {number}. {name}."
    CLINICAL_SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
    CLINICAL_SETTINGS_FILE.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")


PATHOGEN_STATUS_OPTIONS = [
    ("no_detectado", "NO DETECTADO"),
    ("detectado_bajo", "DETECTADO BAJO"),
    ("detectado_medio", "DETECTADO MEDIO"),
    ("detectado_alto", "DETECTADO ALTO"),
]

PATHOGEN_GROUPS = [
    {
        "title": "PATÓGENOS ENTÉRICOS",
        "items": [
            ("shigella", "Shigella spp"),
            ("salmonella", "Salmonella spp"),
            ("campylobacter", "Campylobacter spp"),
            ("rotavirus", "Rotavirus"),
            ("astrovirus", "Astrovirus"),
            ("norovirus", "Norovirus"),
        ],
    },
    {
        "title": "PATÓGENO GÁSTRICO",
        "items": [
            ("helicobacter", "Helicobacter pylori"),
        ],
    },
]

DEFAULT_PATHOGEN_SETTINGS = {
    "shigella": "no_detectado",
    "salmonella": "no_detectado",
    "campylobacter": "no_detectado",
    "rotavirus": "no_detectado",
    "astrovirus": "no_detectado",
    "norovirus": "no_detectado",
    "helicobacter": "detectado_medio",
}


def pathogen_status_label(value: Any) -> str:
    lookup = dict(PATHOGEN_STATUS_OPTIONS)
    return lookup.get(clean_text(value), "NO DETECTADO")


def load_pathogen_settings() -> dict[str, str]:
    if PATHOGEN_SETTINGS_FILE.exists():
        try:
            data = json.loads(PATHOGEN_SETTINGS_FILE.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                merged = dict(DEFAULT_PATHOGEN_SETTINGS)
                merged.update({key: clean_text(value) for key, value in data.items()})
                return merged
        except Exception:
            pass
    PATHOGEN_SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
    PATHOGEN_SETTINGS_FILE.write_text(json.dumps(DEFAULT_PATHOGEN_SETTINGS, ensure_ascii=False, indent=2), encoding="utf-8")
    return dict(DEFAULT_PATHOGEN_SETTINGS)


def save_pathogen_settings(settings: dict[str, str]) -> None:
    valid = {key for key, _ in PATHOGEN_STATUS_OPTIONS}
    current = load_pathogen_settings()
    for key in DEFAULT_PATHOGEN_SETTINGS:
        value = clean_text(settings.get(key))
        current[key] = value if value in valid else "no_detectado"
    PATHOGEN_SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
    PATHOGEN_SETTINGS_FILE.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")


TABLE_HEADER_SPECS = [
    ("ACTINOBACTERIA", "", "30-60%"),
    ("FIRMICUTES", "", "20-40%"),
    ("BACTEROIDETES", "", "30-50%"),
    ("PROTEOBACTERIAS", "", "<1%"),
    ("MICROBIOTA PROTECTORA", "", "30-60%"),
    ("BACTERIAS INMUNOESTIMULANTES (BENEFICIOSAS)", "", "40-70%"),
    ("B. PRODUCTORAS/FACILITADORA DE MUCINA", "2,4%", "1-4% / 40-70%"),
    ("BACTERIAS FORMADORAS DE BUTIRATO", "68,3%", "40-60%"),
    ("PATOGENOS OPORTUNISTAS", "", ""),
    ("M. PROINFLAMATORIA (POTENCIALMENTE PERJUDICIAL)", "21,3%", "<20%"),
    ("M.PRODUCTORA DE H2S", "0,7%", "0-1%"),
    ("M. PRODUCTORA DE HISTAMINA", "0,4%", "<1%"),
    ("M. PRODUCTORA DE TMA Y TMAO", "1,0%", "<2%"),
    ("M. PRODUCTORES DE AMONIACO", "85,5%", "<70%"),
    ("M. PRODUCTORA DE FENOLES", "85,5%", "<60%"),
    ("M.PRODUCTORES DE ACIDOS BILIARES SECUNDARIOS", "85,5%", "<75%"),
    ("M.PRODUCTORA DE SULFATO DE INDOXILO", "85,7%", "<60%"),
    ("MARCADORES DE PATOGENICIDAD Y RESISTENCIA", "NO HAY SOBRECRECIMIENTO NI RESISTENCIA", ""),
]


def table_header_key(title: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", norm(clean_text(title))).strip("_")


def default_table_header_settings() -> dict[str, dict[str, str]]:
    return {table_header_key(title): {"patient": "", "reference": ""} for title, _, _ in TABLE_HEADER_SPECS}


def load_table_header_settings() -> dict[str, dict[str, str]]:
    defaults = default_table_header_settings()
    if TABLE_HEADER_SETTINGS_FILE.exists():
        try:
            data = json.loads(TABLE_HEADER_SETTINGS_FILE.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                for key, value in data.items():
                    if key in defaults and isinstance(value, dict):
                        defaults[key] = {
                            "patient": clean_text(value.get("patient")),
                            "reference": clean_text(value.get("reference")),
                        }
                return defaults
        except Exception:
            pass
    TABLE_HEADER_SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
    TABLE_HEADER_SETTINGS_FILE.write_text(json.dumps(defaults, ensure_ascii=False, indent=2), encoding="utf-8")
    return defaults


def save_table_header_settings(settings: dict[str, dict[str, str]]) -> None:
    current = load_table_header_settings()
    for key in current:
        value = settings.get(key, {})
        current[key] = {
            "patient": clean_text(value.get("patient") if isinstance(value, dict) else ""),
            "reference": clean_text(value.get("reference") if isinstance(value, dict) else ""),
        }
    TABLE_HEADER_SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
    TABLE_HEADER_SETTINGS_FILE.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")


def apply_table_header_settings(groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    settings = load_table_header_settings()
    for group in groups:
        values = settings.get(table_header_key(group.get("title")))
        if not values:
            continue
        patient = clean_text(values.get("patient"))
        reference = clean_text(values.get("reference"))
        if patient:
            group["patient"] = patient
        if reference:
            group["reference"] = reference
    return groups


def session_pdf_path() -> Path | None:
    value = clean_text(load_session_state().get("pdf"))
    return Path(value) if value else None


def session_xlsx_path() -> Path | None:
    value = clean_text(load_session_state().get("xlsx"))
    return Path(value) if value else None


def session_patient_overrides() -> dict[str, str]:
    value = load_session_state().get("patient_overrides", {})
    return value if isinstance(value, dict) else {}


def default_pdf_path() -> Path:
    saved = session_pdf_path()
    if saved and saved.exists():
        return saved
    for folder in (UPLOADS, APP_DIR / "uploads"):
        if folder.exists():
            pdfs = sorted(folder.glob("*.pdf"))
            if pdfs:
                return pdfs[0]
    return DEFAULT_PDF


def default_xlsx_path() -> Path:
    return session_xlsx_path() or Path(MANUAL_SOURCE)


def catalog_to_tsv(catalog: dict[str, list[dict[str, Any]]]) -> str:
    rows = catalog.get("Resultados", [])
    headers = [
        "NAME OF RESEARCH",
        "category",
        "category_pg1",
        "reference",
        "reference_pct",
        "report_tables",
        "display_order",
    ]
    source_keys = {
        "NAME OF RESEARCH": ("NAME OF RESEARCH",),
        "category": ("category", "Category"),
        "category_pg1": ("category_pg1", "Categoria Asignada Pg1"),
        "reference": ("reference", "Reference interval"),
        "reference_pct": ("reference_pct", "Reference interval %"),
        "report_tables": ("report_tables", "Report tables", "tables"),
        "display_order": ("display_order", "Display order", "order"),
    }
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=headers, delimiter="\t", lineterminator="\n", extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({
            header: next((clean_text(row.get(key)) for key in source_keys[header] if clean_text(row.get(key))), "")
            for header in headers
        })
    return buffer.getvalue()


def default_manual_tsv() -> str:
    if MANUAL_CATALOG.exists():
        return MANUAL_CATALOG.read_text(encoding="utf-8")
    if DEFAULT_XLSX.exists():
        text = catalog_to_tsv(load_catalog(DEFAULT_XLSX))
        MANUAL_CATALOG.write_text(text, encoding="utf-8")
        return text
    return "NAME OF RESEARCH\tcategory\tcategory_pg1\treference\treference_pct\treport_tables\tdisplay_order\n"


def load_manual_catalog() -> dict[str, list[dict[str, Any]]]:
    text = default_manual_tsv()
    reader = csv.DictReader(io.StringIO(text), delimiter="\t")
    records: list[dict[str, Any]] = []
    for row in reader:
        record = {clean_text(k): clean_text(v) for k, v in row.items() if k}
        if record.get("NAME OF RESEARCH"):
            record["Category"] = record.get("Category") or record.get("category", "")
            record["Categoria Asignada Pg1"] = record.get("Categoria Asignada Pg1") or record.get("category_pg1", "")
            record["Reference interval"] = record.get("Reference interval") or record.get("reference", "")
            record["Reference interval %"] = record.get("Reference interval %") or record.get("reference_pct", "")
            record["Report tables"] = record.get("Report tables") or record.get("report_tables", "") or record.get("tables", "") or default_report_tables(record.get("NAME OF RESEARCH"))
            record["Display order"] = record.get("Display order") or record.get("display_order", "") or record.get("order", "") or default_display_order(record.get("NAME OF RESEARCH"))
            records.append(record)
    return {"Resultados": records}


CATALOG_FORM_FIELDS = [
    ("NAME OF RESEARCH", "Nombre"),
    ("category", "Grupo del informe"),
    ("category_pg1", "Filo / resumen"),
    ("reference", "Referencia absoluta"),
    ("reference_pct", "Referencia relativa %"),
    ("report_tables", "Fichas/tablas"),
    ("display_order", "Orden"),
]

FUNCTIONAL_CATEGORY_OPTIONS = [
    ("", "Sin grupo"),
    ("1.Actinobacteria", "Actinobacteria"),
    ("2.Firmicutes", "Firmicutes"),
    ("3.Bacteroidetes", "Bacteroidetes"),
    ("4.Other bacteria", "Other bacteria"),
    ("5.OPPORTUNISTIC PATHOGENS", "Opportunistic pathogens"),
    ("6.MARKERS OF PATHOGENICITY AND RESISTANCE", "Pathogenicity / resistance markers"),
    ("7.YEAST FUNGI", "Yeast fungi"),
]

PHYLA_CATEGORY_OPTIONS = [
    ("", "No suma en resumen"),
    ("1.Actinobacteria", "Actinobacteria"),
    ("2.Firmicutes", "Firmicutes"),
    ("3.Bacteroidetes", "Bacteroidetes"),
    ("4.Fusobacteria", "Fusobacteria"),
    ("5.Verrumicrobia", "Verrumicrobia"),
    ("6.Proteo Bacterias", "Proteobacteria"),
    ("7.Otros", "Otros"),
]

REPORT_TABLE_OPTIONS = [
    "ACTINOBACTERIA",
    "FIRMICUTES",
    "BACTEROIDETES",
    "FUSOBACTERIA",
    "VERRUMICROBIA",
    "PROTEOBACTERIAS",
    "MICROBIOTA PROTECTORA",
    "BACTERIAS INMUNOESTIMULANTES (BENEFICIOSAS)",
    "B. PRODUCTORAS/FACILITADORA DE MUCINA",
    "BACTERIAS FORMADORAS DE BUTIRATO",
    "PATOGENOS OPORTUNISTAS",
    "M. PROINFLAMATORIA (POTENCIALMENTE PERJUDICIAL)",
    "M.PRODUCTORA DE H2S",
    "M. PRODUCTORA DE HISTAMINA",
    "M. PRODUCTORA DE TMA Y TMAO",
    "M. PRODUCTORES DE AMONIACO",
    "M. PRODUCTORA DE FENOLES",
    "M.PRODUCTORES DE ACIDOS BILIARES SECUNDARIOS",
    "M.PRODUCTORA DE SULFATO DE INDOXILO",
    "MARCADORES DE PATOGENICIDAD Y RESISTENCIA",
    "HONGOS Y LEVADURAS",
    "ARQUEAS",
]

DEFAULT_REPORT_TABLE_ROWS = {
    "ACTINOBACTERIA": [
        "Bifidobacterium adolescentis", "Bifidobacterium animalis subsp. lactis", "Bifidobacterium bifidum",
        "Bifidobacterium breve", "Bifidobacterium catenulatum ssp", "Bifidobacterium dentium",
        "Bifidobacterium longum subsp. infantis", "Bifidobacterium longum subsp. longum",
        "Bifidobacterium spp", "Coriobacteriia", 'Metabolically active "adult" bifidobacteria **',
        'Metabolically active "infant" bifidobacteria **', "Metabolically active bifidobacteria species, proportion",
    ],
    "FIRMICUTES": [
        "Clostridium difficile gr", "Clostridium leptum gr", "Dialister+Allisonella+Megasphaera+Veillonella",
        "Enterococcus spp", "Erysipelotrichaceae", "Faecalibacterium prausnitzii", "Lachnospiraceae",
        "Lactobacillaceae", "Lactococcus lactis", "Peptoniphilaceae", "Streptococcus spp",
    ],
    "BACTEROIDETES": ["Prevotella spp", "Parabacteroides spp", "Butyricimonas spp", "Bacteroides spp", "Alistipes spp"],
    "FUSOBACTERIA": ["Fusobacteriaceae"],
    "VERRUMICROBIA": ["Akkermansia muciniphila"],
    "PROTEOBACTERIAS": ["Desulfovibrio spp", "E.coli", "Enterobacterales", "Pseudomonas spp"],
    "MICROBIOTA PROTECTORA": [
        "Bifidobacterium adolescentis", "Bifidobacterium breve", "Bifidobacterium catenulatum ssp",
        "Bifidobacterium dentium", "Bifidobacterium longum subsp. infantis", "Faecalibacterium prausnitzii",
        "Akkermansia muciniphila", "Lactobacillaceae", "Bifidobacterium spp", "Clostridium leptum gr", "Lachnospiraceae",
    ],
    "BACTERIAS INMUNOESTIMULANTES (BENEFICIOSAS)": ["Lactobacillaceae", "Akkermansia muciniphila", "Bifidobacterium spp", "Faecalibacterium prausnitzii", "Clostridium leptum gr"],
    "B. PRODUCTORAS/FACILITADORA DE MUCINA": ["Akkermansia muciniphila", "Bacteroides spp", "Faecalibacterium prausnitzii", "Lachnospiraceae", "Lactobacillaceae", "Parabacteroides spp"],
    "BACTERIAS FORMADORAS DE BUTIRATO": ["Lactobacillaceae", "Butyricimonas spp", "Faecalibacterium prausnitzii", "Clostridium leptum gr", "Lachnospiraceae"],
    "PATOGENOS OPORTUNISTAS": ["Pseudomonas spp", "Staphylococcus spp", "Clostridium difficile gr", "Clostridium perfringens gr", "Desulfovibrio spp", "Enterococcus spp", "Fusobacteriaceae", "Peptoniphilaceae", "E.coli", "Enterobacterales", "Erysipelotrichaceae"],
    "M. PROINFLAMATORIA (POTENCIALMENTE PERJUDICIAL)": ["Bacteroides spp", "Clostridium perfringens gr", "Desulfovibrio spp", "E.coli", "Enterobacterales", "Enterococcus spp", "Pseudomonas spp", "Staphylococcus aureus", "Staphylococcus spp"],
    "M.PRODUCTORA DE H2S": ["Clostridium perfringens gr", "Desulfovibrio spp", "Enterobacterales", "Erysipelotrichaceae"],
    "M. PRODUCTORA DE HISTAMINA": ["Clostridium difficile gr", "Clostridium perfringens gr", "E.coli", "Enterobacterales", "Enterococcus spp", "Lactobacillaceae", "Pseudomonas spp", "Staphylococcus aureus", "Staphylococcus spp"],
    "M. PRODUCTORA DE TMA Y TMAO": ["Pseudomonas spp", "Clostridium perfringens gr", "Staphylococcus spp", "E.coli", "Enterobacterales", "Desulfovibrio spp", "Bacteroides spp", "Clostridium leptum gr", "Alistipes spp"],
    "M. PRODUCTORES DE AMONIACO": ["Pseudomonas spp", "Staphylococcus spp", "Clostridium perfringens gr", "Desulfovibrio spp", "Enterococcus spp", "E.coli", "Alistipes spp", "Enterobacterales"],
    "M. PRODUCTORA DE FENOLES": ["Alistipes spp", "Bacteroides spp", "Clostridium perfringens gr", "Desulfovibrio spp", "E.coli", "Enterobacterales", "Enterococcus spp", "Fusobacteriaceae", "Peptoniphilaceae", "Pseudomonas spp", "Staphylococcus spp"],
    "M.PRODUCTORES DE ACIDOS BILIARES SECUNDARIOS": ["Alistipes spp", "Bacteroides spp", "Clostridium leptum gr", "Desulfovibrio spp", "Lachnospiraceae"],
    "M.PRODUCTORA DE SULFATO DE INDOXILO": ["Alistipes spp", "Bacteroides spp", "Clostridium perfringens gr", "Desulfovibrio spp", "E.coli", "Enterobacterales", "Peptoniphilaceae", "Pseudomonas spp"],
    "MARCADORES DE PATOGENICIDAD Y RESISTENCIA": ["Clostridioides difficile", "mecA", "tcdA tcdB", "srr2", "Streptococcus agalactiae", "Staphylococcus aureus", "Pseudomonas spp", "Enterococcus spp"],
    "HONGOS Y LEVADURAS": ["C.albicans", "Candida spp"],
    "ARQUEAS": ["Methanobrevibacter spp"],
}


def default_report_tables(name: Any) -> str:
    name_key = norm(name)
    tables = [table for table, names in DEFAULT_REPORT_TABLE_ROWS.items() if any(norm(row_name) == name_key for row_name in names)]
    return ", ".join(tables)


def default_display_order(name: Any) -> str:
    name_key = norm(name)
    positions = []
    for names in DEFAULT_REPORT_TABLE_ROWS.values():
        for index, row_name in enumerate(names, start=1):
            if norm(row_name) == name_key:
                positions.append(index * 10)
    return str(min(positions)) if positions else ""


def catalog_records_to_tsv(records: list[dict[str, Any]]) -> str:
    buffer = io.StringIO()
    headers = [field for field, _ in CATALOG_FORM_FIELDS]
    writer = csv.DictWriter(buffer, fieldnames=headers, delimiter="\t", lineterminator="\n")
    writer.writeheader()
    for row in records:
        name = clean_text(row.get("NAME OF RESEARCH"))
        if not name:
            continue
        writer.writerow({field: clean_text(row.get(field)) for field, _ in CATALOG_FORM_FIELDS})
    return buffer.getvalue()


DEFAULT_CONSIDERATIONS = """Consideraciones:

DISBIOSIS INTESTINAL MODERADA-SEVERA.

1- Biomasa bacteriana total disminuida con baja diversidad.
2- Perfil metabólico mixto Firmicutes/ butirato dominante con componente bilio-tolerante alto. Enterotipo 3.
3- Déficit severo de simbiontes protectores clásicos asociado a baja resiliencia.
4- Aumento severo de permeabilidad intestinal.
5- Sin evidencia de Inflamación intestinal activa alta.
6- Eje mucina severamente debilitada, compatible con fragilidad funcional de la mucosa.
7- Soporte trófico colónico conservado.
8- Sobrecrecimiento micótico moderado.
9- Carga baja de oportunistas bacterianos. Sin evidencia de infección toxigénica activa por C. difficile.
10- Sin evidencia de componente infeccioso entérico.
11- Evidencia de componente gastroduodenal: H. pylori Detectado medio.
12- Marcador de resistencia (mecA) no detectado.

CONSIDERACIONES NUTRICIONALES GENERALES:

SEMANA 1 a la 2. Control de inflamación.
ELIMINAR: Embutidos, carnes rojas, huevos y ultraprocesados, lácteos, alcohol, café, laxantes, vinagre y fermentados. Picantes, proteína en polvo, edulcorantes, miel, panela, jugos y gaseosas. Arroz, harinas blancas, cereales refinados. Ajo, cebolla, puerro, brócoli, coliflor, manzana, pera.

INCLUIR.
Proteínas magras: pescado, pollo y pavo.
Verduras cocidas: calabacín, berenjena, auyama, acelgas, espinaca, hinojo, zanahoria. Batata, yuca y plátano verde en porciones moderadas.
Aceite de oliva y aguacate. Té de manzanilla y té verde suave, cacao puro, arándanos. Infusiones de jengibre suave. 1 taza de caldo al día. Frutas: kiwi, lechosa.

PROBIÓTICOS:
* Saccharomyces bourlardii por 4 semanas. Descansar un mes. Repetir durante dos meses solo dos semanas por mes.
* Candida support complex. Lactobacillus reuteri 8 semanas.

SEMANA 3 a la 4. Recolonización
Reintroducir fibra gradualmente: un alimento nuevo cada tres días. Incluir frutas: kiwi, fresas, lechosa. Ajo y cebolla en micro dosis para evaluar tolerancia. Lactobacillus rhamnosus, L. plantarum. Bifidobacterium longum, B. bifidum, B. lactis. Akkermansia.
Probar tolerancia al aumentar progresivamente variedad vegetal. Té verde suave y caldos de hueso.

SEMANA 5-8 Consolidación
Proteína animal moderada y priorizar pescado. Rotar variedad de vegetales por semana y retirar el que empeore síntomas. Incluir granada, fresas, arándanos, manzana, gelatina natural, caldo de huesos, cúrcuma, jengibre suave. Mantenimiento de ciclos de 2 meses con Bifidos + Lactobacillus (8-12 cepas). Descansar un mes. Rotación de probióticos multi cepas. Cuidar la hidratación.

OBJETIVOS:
• Fortalecer la barrera intestinal. Corregir estreñimiento.
• Mejorar el soporte trófico colónico. Mejorar sesgo bilio-tolerante y reponer simbiontes claves ausentes.
• Control micótico. Evaluar marcadores inflamatorios en la semana 8.

Tomar el probiótico con bebidas frescas. Evitar ingerirlos con comidas o bebidas calientes.
"""

YEAST_NAMES = [
    "Candida albicans",
    "Candida auris",
    "Candida glabrata",
    "Candida tropicalis",
    "Clavispora lusitaniae (Candida lusitaniae)",
    "Debaryomyces hansenii (C.famata)",
    "Kluyveromyces marxianus (C.kefyr)",
    "Malassezia furfur",
    "Malassezia spp.",
    "Meyerozyma guilliermondii (C.guilliermondi)",
    "Pichia kudriavzevii (C.krusei)",
    "Saccharomyces cerevisiae",
]

YEAST_STATUSES = [
    ("no_evaluado", "No evaluado"),
    ("debil", "Debil"),
    ("moderada", "Moderada"),
    ("elevado", "Elevado"),
]

YEAST_STATUS_POSITIONS = {
    "no_evaluado": 0.0,
    "debil": 15.0,
    "moderada": 50.0,
    "elevado": 86.0,
}


def load_considerations() -> str:
    if CONSIDERATIONS_FILE.exists():
        text = CONSIDERATIONS_FILE.read_text(encoding="utf-8")
        return text if clean_text(text) else DEFAULT_CONSIDERATIONS
    CONSIDERATIONS_FILE.write_text(DEFAULT_CONSIDERATIONS, encoding="utf-8")
    return DEFAULT_CONSIDERATIONS


def default_yeast_records() -> list[dict[str, str]]:
    return [{"name": name, "status": "no_evaluado", "value": "", "notes": ""} for name in YEAST_NAMES]


def save_yeast_profile(records: list[dict[str, Any]]) -> None:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=["name", "status", "value", "notes"], delimiter="\t", lineterminator="\n")
    writer.writeheader()
    for record in records:
        name = clean_text(record.get("name"))
        if not name:
            continue
        status = clean_text(record.get("status")) or "no_evaluado"
        if status not in {key for key, _ in YEAST_STATUSES}:
            status = "no_evaluado"
        writer.writerow({
            "name": name,
            "status": status,
            "value": clean_text(record.get("value")),
            "notes": clean_text(record.get("notes")),
        })
    YEAST_PROFILE_FILE.write_text(buffer.getvalue(), encoding="utf-8")


def load_yeast_profile() -> list[dict[str, str]]:
    if not YEAST_PROFILE_FILE.exists():
        records = default_yeast_records()
        save_yeast_profile(records)
        return records
    reader = csv.DictReader(io.StringIO(YEAST_PROFILE_FILE.read_text(encoding="utf-8")), delimiter="\t")
    by_name: dict[str, dict[str, str]] = {}
    for row in reader:
        name = clean_text(row.get("name"))
        if not name:
            continue
        status = clean_text(row.get("status")) or "no_evaluado"
        if status not in {key for key, _ in YEAST_STATUSES}:
            status = "no_evaluado"
        by_name[norm(name)] = {
            "name": name,
            "status": status,
            "value": clean_text(row.get("value")),
            "notes": clean_text(row.get("notes")),
        }
    records = []
    for name in YEAST_NAMES:
        records.append(by_name.get(norm(name), {"name": name, "status": "no_evaluado", "value": "", "notes": ""}))
    return records


def yeast_status_label(status: str) -> str:
    return next((label for key, label in YEAST_STATUSES if key == status), "No evaluado")


def yeast_marker(record: dict[str, Any]) -> float:
    numeric = metric_float(record.get("value"))
    if numeric is not None:
        return max(0.0, min(numeric, 100.0))
    return YEAST_STATUS_POSITIONS.get(clean_text(record.get("status")), 0.0)


def free_text_to_html(text: str) -> str:
    blocks = [block.strip() for block in re.split(r"\n\s*\n", text.strip()) if block.strip()]
    if not blocks:
        return "<p><b>Consideraciones:</b></p>"
    html_blocks: list[str] = []
    for block in blocks:
        safe = "<br>".join(h(line.rstrip()) for line in block.splitlines())
        if ":" in block.splitlines()[0] and len(block.splitlines()[0]) <= 90:
            first, *rest = safe.split("<br>")
            safe = f"<b>{first}</b>" + (("<br>" + "<br>".join(rest)) if rest else "")
        html_blocks.append(f"<p>{safe}</p>")
    return "\n".join(html_blocks)


def lines_to_html(text: Any) -> str:
    lines = [line.rstrip() for line in str(text or "").replace("\r\n", "\n").splitlines() if clean_text(line)]
    return "<br>".join(h(line) for line in lines)


def extract_patient(text: str) -> dict[str, str]:
    patient: dict[str, str] = {}
    line_fields = {
        "FULL NAME": "Nombre",
        "CONTAINER ID": "Contenedor",
        "PHYSICIAN": "Medico",
    }
    for raw, label in line_fields.items():
        match = re.search(rf"{re.escape(raw)}:[ \t]*([^\n\r]*)", text, flags=re.I)
        if match:
            patient[label] = clean_text(match.group(1))
    sex = re.search(r"SEX:\s*(female|male|femenino|masculino|f|m)\b", text, flags=re.I)
    if sex:
        patient["Sexo"] = clean_text(sex.group(1))
    dob = re.search(r"DATE OF BIRTH:\s*(\d{1,2}/\d{1,2}/\d{2,4})", text, flags=re.I)
    if dob:
        patient["Fecha de nacimiento"] = dob.group(1)
    sample = re.search(r"DATE OF SAMPLING:\s*(\d{1,2}/\d{1,2}/\d{2,4})", text, flags=re.I)
    if sample:
        patient["Fecha de muestra"] = sample.group(1)
    # Some PDFs wrap the first header line; keep the most important fallbacks.
    patient.setdefault("Nombre", "")
    patient.setdefault("Sexo", "")
    patient.setdefault("Fecha de nacimiento", "")
    patient.setdefault("Fecha de muestra", "")
    return patient


def words_to_lines(words: list[dict[str, Any]], tolerance: float = 3.0) -> list[list[dict[str, Any]]]:
    lines: list[list[dict[str, Any]]] = []
    for word in sorted(words, key=lambda w: (w["top"], w["x0"])):
        for line in lines:
            if abs(line[0]["top"] - word["top"]) <= tolerance:
                line.append(word)
                break
        else:
            lines.append([word])
    return [sorted(line, key=lambda w: w["x0"]) for line in lines]


def line_text(words: list[dict[str, Any]]) -> str:
    return clean_text(" ".join(w["text"] for w in words))


def is_invalid_text(value: Any) -> bool:
    text = clean_text(value).lower()
    if not text:
        return False
    return bool(re.search(r"\binv(?:a|á)lid(?:o|a)?\b", text))


def display_result_value(value: Any) -> str:
    text = clean_text(value)
    return "no detectado" if is_invalid_text(text) else text


def extract_observation_from_page(page: pdfplumber.page.Page) -> str:
    words = [
        w
        for w in page.extract_words(x_tolerance=1, y_tolerance=3)
        if w["x0"] >= 300 and 100 <= w["top"] <= 215
    ]
    return "\n".join(line_text(line) for line in words_to_lines(words) if line_text(line))


def flatten_page1_tables(page: pdfplumber.page.Page) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    current_group = ""
    for table in page.extract_tables() or []:
        for row in table:
            cells = [clean_text(c) for c in row]
            if not any(cells):
                continue
            name = cells[0] or current_group
            if cells[0] and not cells[1] and not cells[2]:
                current_group = cells[0]
            rows.append(
                {
                    "name": name,
                    "result": display_result_value(cells[1] if len(cells) > 1 else ""),
                    "reference": cells[2] if len(cells) > 2 else "",
                    "unit": cells[3] if len(cells) > 3 else "",
                }
            )
    return rows


def split_result_line(line: str) -> dict[str, str] | None:
    line = clean_text(line)
    match = re.match(
        r"^(?P<name>.+?)\s+(?P<rest>(?:Not assessed at this age|not detected|detected|inv(?:a|á)lid(?:o|a)?)(?:\s+.*)?|[><≥≤]?\s*-?\d.+|-.*)$",
        line,
        flags=re.I,
    )
    if not match:
        return None
    name = clean_text(match.group("name"))
    rest = clean_text(match.group("rest"))
    result = ""
    remaining = ""
    for phrase in ["Not assessed at this age", "not detected", "detected", "invalid", "invalido", "invalida", "inválido", "inválida"]:
        if rest.lower().startswith(phrase.lower()):
            result = phrase
            remaining = clean_text(rest[len(phrase) :])
            break
    if not result:
        token = re.match(r"([><≥≤]?\s*-?\d+(?:[.,]\d+)?\*?(?:\s*[↘↗])?|-)", rest)
        if not token:
            return None
        result = clean_text(token.group(1))
        remaining = clean_text(rest[token.end() :])

    unit = ""
    unit_match = re.search(r"\s(%|PCS|Lg\s*\(.+?\))$", remaining)
    if unit_match:
        unit = clean_text(unit_match.group(1))
        remaining = clean_text(remaining[: unit_match.start()])
    for phrase in ["not detected", "detected"]:
        suffix = f" {phrase}"
        if name.lower().endswith(suffix) and result.lower() == phrase:
            name = clean_text(name[: -len(suffix)])
            remaining = phrase if not remaining else remaining
    return {"name": name, "result": display_result_value(result), "reference": remaining, "unit": unit}


def parse_page1_text(page1_text: str) -> list[dict[str, str]]:
    lines = [clean_text(line) for line in page1_text.splitlines()]
    rows: list[dict[str, str]] = []
    in_table = False
    current_group = ""
    skip = {
        "REFERENCE UNIT",
        "NAME OF RESEARCH RESULT",
        "INTERVAL MEASUREMENTS",
        "Research result",
    }
    for line in lines:
        if "NAME OF RESEARCH RESULT" in line:
            in_table = True
            continue
        if line.startswith("Assay date:"):
            break
        if not in_table or not line or line in skip:
            continue
        parsed = split_result_line(line)
        if parsed:
            if parsed["name"].lower() in {"proportion", "diversity, number of taxa", "total amount"} and current_group:
                parsed["name"] = f"{current_group} - {parsed['name']}"
            rows.append(parsed)
        else:
            current_group = line.rstrip(".")
    return rows


NUM_RE = re.compile(r"(?:[<>>=≤≥]\s*)?-?\d+(?:[.,]\d+)?|-|(?:not|no)\s+detect(?:ed|ado|ada|d)?|detected|inv(?:a|á)lid(?:o|a)?", re.I)


def number_or_none(value: str) -> float | None:
    value = clean_text(value).replace(",", ".")
    value = value.replace(">", "").replace("<", "").replace("≥", "").replace("≤", "")
    value = value.replace("*", "").replace("↘", "").replace("↗", "").strip()
    if value == "-":
        return 0.0
    try:
        return float(value)
    except ValueError:
        return None


def first_number(text: str) -> float | None:
    match = NUM_RE.search(clean_text(text))
    if not match:
        return None
    return number_or_none(match.group(0))


def parse_numeric_row(
    name: str,
    rest: str,
    catalog_row: dict[str, Any] | None = None,
    *,
    absolute_text: str = "",
    absolute_reference: str = "",
    relative_text: str = "",
    relative_reference: str = "",
) -> dict[str, Any]:
    absolute_display = display_result_value(absolute_text)
    relative_display = display_result_value(relative_text)
    value = first_number(absolute_text or rest)
    relative = first_number(relative_text)
    if relative is not None:
        relative = relative / 100
    return {
        "name": name,
        "raw": rest,
        "value": value,
        "relative": relative,
        "absolute_text": clean_text(absolute_display),
        "absolute_reference": clean_text(absolute_reference),
        "relative_text": clean_text(relative_display),
        "relative_reference": clean_text(relative_reference),
        "category": clean_text((catalog_row or {}).get("Category")),
        "category_pg1": clean_text((catalog_row or {}).get("Categoria Asignada Pg1")),
        "reference": clean_text((catalog_row or {}).get("Reference interval")),
        "reference_pct": clean_text((catalog_row or {}).get("Reference interval %")),
        "report_tables": clean_text((catalog_row or {}).get("Report tables")),
        "display_order": clean_text((catalog_row or {}).get("Display order")),
    }


def extract_page2_rows_from_page(page: pdfplumber.page.Page, catalog: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    expected = []
    for row in catalog.get("Resultados", []):
        name = clean_text(row.get("NAME OF RESEARCH"))
        if name:
            expected.append((norm(name), name, row))
    expected.sort(key=lambda x: len(x[0]), reverse=True)

    rows_by_name: dict[str, dict[str, Any]] = {}
    words = page.extract_words(x_tolerance=1, y_tolerance=3)
    for line_words in words_to_lines(words):
        if not line_words:
            continue
        line_clean = line_text(line_words)
        line_name = line_text([w for w in line_words if w["x0"] < 270])
        line_norm = norm(line_name)
        if not line_norm:
            continue
        for expected_norm, expected_name, catalog_row in expected:
            if line_norm.startswith(expected_norm):
                absolute = line_text([w for w in line_words if 270 <= w["x0"] < 335])
                absolute_ref = line_text([w for w in line_words if 335 <= w["x0"] < 420])
                relative = line_text([w for w in line_words if 420 <= w["x0"] < 480])
                relative_ref = line_text([w for w in line_words if w["x0"] >= 480])
                rest = line_clean.replace(line_name, "", 1).strip(" :-")
                rows_by_name[expected_name] = parse_numeric_row(
                    expected_name,
                    rest,
                    catalog_row,
                    absolute_text=absolute,
                    absolute_reference=absolute_ref,
                    relative_text=relative,
                    relative_reference=relative_ref,
                )
                break
    return list(rows_by_name.values())


def row_value(rows: list[dict[str, Any]], name: str, key: str = "relative") -> float | None:
    name_n = norm(name)
    for row in rows:
        if norm(row["name"]) == name_n:
            value = row.get(key)
            return value if isinstance(value, (int, float)) else None
    return None


FIRMICUTES_NAMES = [
    "Clostridium difficile gr",
    "Clostridium leptum gr",
    "Dialister+Allisonella+Megasphaera+Veillonella",
    "Enterococcus spp",
    "Erysipelotrichaceae",
    "Faecalibacterium prausnitzii",
    "Lachnospiraceae",
    "Lactobacillaceae",
    "Lactococcus lactis",
    "Peptoniphilaceae",
    "Streptococcus spp",
]

BACTEROIDETES_NAMES = [
    "Alistipes spp",
    "Bacteroides spp",
    "Butyricimonas spp",
    "Parabacteroides spp",
    "Prevotella spp",
]


def named_total(rows: list[dict[str, Any]], names: list[str], key: str = "relative") -> float:
    return sum(row_value(rows, name, key) or 0 for name in names)


def calculate_metrics(rows: list[dict[str, Any]]) -> dict[str, str]:
    by_category: dict[str, float] = {}
    for row in rows:
        cat = row.get("category")
        rel = row.get("relative")
        if cat and isinstance(rel, (int, float)):
            by_category[cat] = by_category.get(cat, 0.0) + rel

    firm = phylum_total(rows, "2.Firmicutes") or by_category.get("2.Firmicutes") or named_total(rows, FIRMICUTES_NAMES)
    bact = phylum_total(rows, "3.Bacteroidetes") or by_category.get("3.Bacteroidetes") or named_total(rows, BACTEROIDETES_NAMES)
    bifido = row_value(rows, "Bifidobacterium spp", "value")
    entero = row_value(rows, "Enterobacterales", "value")
    prev = row_value(rows, "Prevotella spp", "relative")
    bacteroides = row_value(rows, "Bacteroides spp", "relative")

    def div(a: float | None, b: float | None) -> str:
        if a is None or b in (None, 0):
            return "-"
        return f"{a / b:.2f}"

    actino_proteo = "-"
    if bifido is not None and entero is not None:
        actino_proteo = f"{(math.pow(10, bifido) / math.pow(10, entero)):.2f}"

    return {
        "firm_bact": div(firm, bact),
        "actino_proteo": actino_proteo,
        "prev_bacteroides": div(prev, bacteroides),
        "diversity": clean_text(next((r["raw"] for r in rows if norm(r["name"]) == norm("Diversity, number of taxa:")), "")),
        "normal_microbiota": fmt_number(row_value(rows, "NORMAL MICROBIOTA", "relative"), percent=True),
    }


def fmt_number(value: float | None, *, percent: bool = False) -> str:
    if value is None:
        return "-"
    if percent:
        return f"{value * 100:.2f} %".replace(".", ",")
    return f"{value:.2f}".replace(".", ",")


def page2_row(rows: list[dict[str, Any]], name: str) -> dict[str, Any]:
    return next((row for row in rows if norm(row["name"]) == norm(name)), {"value": None, "relative": None, "raw": ""})


def page1_lookup(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    aliases = {
        "Total bacterial mass (TBM)": ["Total bacterial mass (TBM)"],
        "diversity, number of taxa": ["Normal microbiota - diversity, number of taxa", "diversity, number of taxa"],
        "proportion": ["Normal microbiota - proportion", "proportion"],
        "total amount": ["Bifidobacterium spp - total amount", "total amount"],
        'metabolically active "child" species, diversity': [
            "metabolically active “child” species, diversity",
            'metabolically active "child" species, diversity',
        ],
        'metabolically active "child" species, proportion**': [
            "metabolically active “child” species, proportion**",
            'metabolically active "child" species, proportion**',
        ],
        "metabolically active species, proportion": ["metabolically active species, proportion"],
        "Lactobacillaceae, quantity": ["Lactobacillaceae, quantity"],
        "Bacteroidetes, presence": ["Bacteroidetes, presence"],
        "Firmicutes/Bacteroidetes, ratio:": ["Firmicutes/Bacteroidetes, ratio:"],
        "Yeast fungi, quantity": ["Yeast fungi, quantity"],
        "Opportunistic pathogens, proportion": ["Opportunistic pathogens, proportion"],
        "Candida albicans": ["Candida albicans"],
        "Clostridioides difficile": ["Clostridioides difficile"],
        "mecA": ["mecA"],
        "pathogenic agents Enterobacterales": ["pathogenic agents Enterobacterales"],
        "Staphylococcus aureus": ["Staphylococcus aureus"],
        "Streptococcus agalactiae with srr2": ["Streptococcus agalactiae with srr2"],
        "tcdA, tcdB": ["tcdA, tcdB"],
    }
    by_norm = {norm(row["name"]): row for row in rows}
    output: dict[str, dict[str, str]] = {}
    for target, candidates in aliases.items():
        for candidate in candidates:
            row = by_norm.get(norm(candidate))
            if row:
                output[target] = row
                break
    return output


def final_page1_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    lookup = page1_lookup(rows)

    def group(title: str) -> dict[str, str]:
        return {"kind": "group", "name": title, "result": "", "reference": "", "unit": ""}

    def row(name: str, label: str | None = None, bold: bool = False) -> dict[str, str]:
        source = lookup.get(name, {})
        return {
            "kind": "row bold" if bold else "row",
            "name": label or name,
            "result": display_result_value(source.get("result", "")),
            "reference": source.get("reference", ""),
            "unit": source.get("unit", ""),
        }

    return [
        group("Total bacterial mass (TBM)"),
        row("Total bacterial mass (TBM)"),
        group("Normal microbiota"),
        row("diversity, number of taxa"),
        row("proportion"),
        group("Bifidobacterium spp."),
        row('metabolically active "child" species, diversity'),
        row('metabolically active "child" species, proportion**'),
        row("metabolically active species, proportion"),
        row("total amount"),
        row("Lactobacillaceae, quantity", "Lactobacillaceae, quantity", True),
        row("Bacteroidetes, presence", "Bacteroidetes, presence", True),
        row("Firmicutes/Bacteroidetes, ratio:", "Firmicutes/Bacteroidetes ratio", True),
        row("Yeast fungi, quantity", "Yeast fungi, quantity", True),
        row("Opportunistic pathogens, proportion", "Opportunistic pathogens, proportion", True),
        group("Markers of pathogenicity and resistance, presence"),
        row("Candida albicans"),
        row("Clostridioides difficile"),
        row("mecA"),
        row("pathogenic agents Enterobacterales"),
        row("Staphylococcus aureus"),
        row("Streptococcus agalactiae with srr2"),
        row("tcdA, tcdB"),
    ]


def extract_report(pdf_path: Path, xlsx_path: Path) -> ExtractedReport:
    catalog = load_catalog(xlsx_path)
    with pdfplumber.open(pdf_path) as pdf:
        page_texts = [page.extract_text(x_tolerance=1, y_tolerance=3) or "" for page in pdf.pages]
        page1_text = page_texts[0] if page_texts else ""
        page2_text = page_texts[1] if len(page_texts) > 1 else ""
        page1_rows = parse_page1_text(page1_text)
        if not page1_rows and pdf.pages:
            page1_rows = flatten_page1_tables(pdf.pages[0])
        observation = extract_observation_from_page(pdf.pages[0]) if pdf.pages else ""
        page2_rows = extract_page2_rows_from_page(pdf.pages[1], catalog) if len(pdf.pages) > 1 else []
    return ExtractedReport(
        patient=extract_patient(page1_text + "\n" + page2_text),
        observation=observation,
        page1_rows=page1_rows,
        page2_rows=page2_rows,
        catalog=catalog,
    )


def h(value: Any) -> str:
    return escape(str(value if value is not None else ""), quote=True)


def asset_url(filename: str) -> str:
    return f"/assets/{filename}"


def render_report(pdf: Path, xlsx: Path, patient_overrides: dict[str, str] | None = None) -> str:
    if not pdf.exists():
        return build_upload_required_html(pdf, xlsx)
    data = extract_report(pdf, xlsx)
    metrics = calculate_metrics(data.page2_rows)
    logo = "LOGOTIPO_GENOMA_CON_TEXTO_Mesa13338750993692639.png"
    firma = "FIRMA318127053262214.png"
    fondo = "FONDO08012234735382195.jpg"
    selected_bullets = [
        "Total bacterial mass",
        "NORMAL MICROBIOTA",
        "Bifidobacterium spp",
        "Lactobacillaceae",
        "Bacteroides spp",
        "OPPORTUNISTIC PATHOGENS",
        "Yeast fungi, quantity",
    ]
    bullets = [r for r in data.page2_rows if r["name"] in selected_bullets]
    return build_html(pdf, xlsx, data, metrics, bullets, logo, firma, fondo, patient_overrides or {})


def patient_overrides_from_params(params: dict[str, list[str]]) -> dict[str, str]:
    mapping = {
        "patient_nombre": "Nombre",
        "patient_fecha": "Fecha de muestra",
        "patient_ci": "Cedula",
        "patient_sexo": "Sexo",
        "patient_nacimiento": "Fecha de nacimiento",
    }
    overrides: dict[str, str] = {}
    for param, field in mapping.items():
        if param in params:
            overrides[field] = clean_text(params.get(param, [""])[0])
    return overrides


def merged_patient_overrides(params: dict[str, list[str]]) -> dict[str, str]:
    overrides = session_patient_overrides()
    overrides.update(patient_overrides_from_params(params))
    return overrides


def fmt_pct(value: Any) -> str:
    if isinstance(value, (int, float)):
        return f"{value * 100:.1f}%"
    return "-"


def pct_text(value: float | None, digits: int = 1) -> str:
    if not isinstance(value, (int, float)):
        return "-"
    return f"{value * 100:.{digits}f}%".replace(".", ",")


def fmt_abs(value: Any) -> str:
    if isinstance(value, (int, float)):
        if value == 0:
            return "-"
        return f"{value:.1f}".replace(".", ",")
    return "-"


def fmt_summary_abs(value: Any) -> str:
    if isinstance(value, (int, float)):
        return f"{value:.2f}".replace(".", ",")
    return "0,00"


def row_by_name(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {norm(row["name"]): row for row in rows}


def category_total(rows: list[dict[str, Any]], category: str) -> float:
    return sum(
        float(row["relative"])
        for row in rows
        if row.get("category") == category and isinstance(row.get("relative"), (int, float))
    )


def phylum_total(rows: list[dict[str, Any]], phylum: str) -> float:
    return sum(
        float(row["relative"])
        for row in rows
        if (row.get("category_pg1") or row.get("category")) == phylum and isinstance(row.get("relative"), (int, float))
    )


def absolute_pool(rows: list[dict[str, Any]], names: list[str]) -> float:
    total = 0.0
    for name in names:
        value = row_value(rows, name, "value")
        if isinstance(value, (int, float)) and value > 0:
            total += math.pow(10, value)
    return math.log10(total) if total > 0 else 0.0


def category_absolute_total(rows: list[dict[str, Any]], category: str) -> float:
    return absolute_pool(rows, [row["name"] for row in rows if row.get("category") == category])


def phylum_absolute_total(rows: list[dict[str, Any]], phylum: str) -> float:
    return absolute_pool(rows, [row["name"] for row in rows if (row.get("category_pg1") or row.get("category")) == phylum])


def detail_row(rows_map: dict[str, dict[str, Any]], name: str) -> str:
    row = rows_map.get(norm(name), {"name": name})
    absolute = h(row.get("absolute_text") or fmt_abs(row.get("value")))
    relative = h(row.get("relative_text") or fmt_pct(row.get("relative")))
    return (
        f"<tr><td>{h(row.get('name') or name)}</td>"
        f"<td>{absolute}</td><td>{h(row.get('reference') or row.get('absolute_reference') or '')}</td>"
        f"<td>{relative}</td><td>{h(row.get('reference_pct') or row.get('relative_reference') or '')}</td></tr>"
    )


def detail_group(title: str, patient: str = "", reference: str = "") -> str:
    patient_html = f"<span>*PACIENTE: {h(patient)}</span>" if patient else "<span></span>"
    reference_html = f"<span>*V.R: {h(reference)}</span>" if reference else "<span></span>"
    return f'<tr class="section-row"><td colspan="5"><b>{h(title)}</b>{patient_html}{reference_html}</td></tr>'


def table_key(value: Any) -> str:
    return norm(value).replace(" ", "")


def table_memberships(value: Any) -> set[str]:
    text = clean_text(value)
    if not text:
        return set()
    parts = [part.strip() for part in re.split(r"[|,;]\s*", text) if part.strip()]
    return {table_key(part) for part in parts}


def order_value(value: Any) -> float | None:
    text = clean_text(value).replace(",", ".")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def catalog_map(catalog: dict[str, list[dict[str, Any]]] | None) -> dict[str, dict[str, Any]]:
    return {
        norm(row.get("NAME OF RESEARCH")): row
        for row in (catalog or {}).get("Resultados", [])
        if clean_text(row.get("NAME OF RESEARCH"))
    }


def rows_for_detail_group(
    group: dict[str, Any],
    rows: list[dict[str, Any]],
    catalog: dict[str, list[dict[str, Any]]] | None,
) -> list[str]:
    base_rows = [clean_text(name) for name in group.get("rows", []) if clean_text(name)]
    seen = {norm(name) for name in base_rows}
    key = table_key(group.get("title", ""))
    catalog_rows = catalog_map(catalog)
    row_lookup = row_by_name(rows)

    for catalog_row in catalog_rows.values():
        if key in table_memberships(catalog_row.get("Report tables")):
            name = clean_text(catalog_row.get("NAME OF RESEARCH"))
            if name and norm(name) not in seen:
                base_rows.append(name)
                seen.add(norm(name))

    def sort_key(name: str) -> tuple[int, float, int]:
        row = row_lookup.get(norm(name), {})
        catalog_row = catalog_rows.get(norm(name), {})
        explicit = order_value(row.get("display_order") or catalog_row.get("Display order"))
        fallback = base_rows.index(name)
        return (0 if explicit is not None else 1, explicit if explicit is not None else 0, fallback)

    return sorted(base_rows, key=sort_key)


def detail_table(rows: list[dict[str, Any]], groups: list[dict[str, Any]], *, summary: bool = False, catalog: dict[str, list[dict[str, Any]]] | None = None) -> str:
    rows_map = row_by_name(rows)
    body: list[str] = []
    summary_html = ""
    if summary:
        phyla = [
            ("1.Actinobacteria", "Actinobacteria"),
            ("2.Firmicutes", "Firmicutes"),
            ("3.Bacteroidetes", "Bacteroidetes"),
            ("Fusobacteria", "Fusobacteria"),
            ("Verrucomicrobia", "Verrumicrobia"),
            ("Proteo Bacterias", "Proteobacteria"),
            ("Otros", "Otros"),
        ]
        summary_rows = []
        for index, (category, label) in enumerate(phyla, start=1):
            relative_value = phylum_total(rows, category)
            absolute_value = phylum_absolute_total(rows, category)
            if label == "Fusobacteria":
                relative_value = row_value(rows, "Fusobacteriaceae", "relative") or 0
                absolute_value = absolute_pool(rows, ["Fusobacteriaceae"])
            if label == "Verrumicrobia":
                relative_value = row_value(rows, "Akkermansia muciniphila", "relative") or 0
                absolute_value = absolute_pool(rows, ["Akkermansia muciniphila"])
            if label == "Proteobacteria":
                proteo_names = ["Desulfovibrio spp", "E.coli", "Enterobacterales", "Pseudomonas spp"]
                relative_value = sum(row_value(rows, n, "relative") or 0 for n in proteo_names)
                absolute_value = absolute_pool(rows, proteo_names)
            if label == "Otros":
                relative_value = row_value(rows, "Methanobrevibacter spp", "relative") or 0
                absolute_value = absolute_pool(rows, ["Methanobrevibacter spp"])
            summary_rows.append(f"<tr><td>{h(label)}</td><td>{fmt_summary_abs(absolute_value)}</td><td>{pct_text(relative_value, 2).replace('%', ' %')}</td></tr>")
        summary_html = f"""
          <table class="phyla-summary-table">
            <thead><tr><th>Fila</th><th>RESULT ( ABSOLUTE,<br>Lg (GE/g feces)*)</th><th>RELATIVE, VALUE %</th></tr></thead>
            <tbody>{''.join(summary_rows)}</tbody>
          </table>
        """
    for group in groups:
        body.append(detail_group(group["title"], group.get("patient", ""), group.get("reference", "")))
        body.extend(detail_row(rows_map, name) for name in rows_for_detail_group(group, rows, catalog))
    return f"""
      {summary_html}
      <table class="detail-table">
        <thead><tr><th>NAME OF RESEARCH</th><th>RESULT ( ABSOLUTE,<br>Lg (GE/g feces)*)</th><th>REFERENCE INTERVAL (<br>ABSOLUTE, Lg (GE/g feces)*)</th><th>RELATIVE,<br>VALUE %</th><th>REFERENCE<br>INTERVAL %</th></tr></thead>
        <tbody>{''.join(body)}</tbody>
      </table>
    """


def bar_fill(row: dict[str, Any] | None) -> float:
    value = row.get("relative") if row else None
    if not isinstance(value, (int, float)):
        return 0
    return max(0, min(value * 100, 100))


def parse_pct_reference(reference: Any) -> tuple[float, float]:
    text = clean_text(reference).replace(",", ".")
    nums = [float(match) for match in re.findall(r"\d+(?:\.\d+)?", text)]
    if len(nums) >= 2:
        return max(0, min(nums[0], 100)), max(0, min(nums[1], 100))
    if len(nums) == 1:
        if "<" in text:
            return 0, max(0, min(nums[0], 100))
        if ">" in text:
            return max(0, min(nums[0], 100)), 100
    return 10, 60


def metric_float(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value).replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def metric_card(value: Any, label: str, reference: str, low: float, high: float, *, scale: float | None = None) -> str:
    numeric = metric_float(value)
    ceiling = scale or max(high * 1.35, (numeric or 0) * 1.15, 1)
    marker = max(0, min(((numeric or 0) / ceiling) * 100, 100))
    start = max(0, min((low / ceiling) * 100, 100))
    end = max(start, min((high / ceiling) * 100, 100))
    status = "ok" if numeric is not None and low <= numeric <= high else "warn"
    return f"""
      <div class="metric-card {status}" style="--value:{marker:.2f}%; --low:{start:.2f}%; --high:{end:.2f}%">
        <strong>{h(value)}</strong>
        <div class="metric-gauge"><i></i><b></b></div>
        <p>{h(label)}<br><span>{h(reference)}</span></p>
      </div>
    """


def comparison_card(
    ratio: Any,
    title: str,
    reference: str,
    first_label: str,
    first_value: float | None,
    second_label: str,
    second_value: float | None,
    *,
    value_kind: str = "percent",
    scale: float | None = None,
) -> str:
    first = float(first_value or 0)
    second = float(second_value or 0)
    ceiling = scale or max(first, second, 1)
    first_width = max(1.5, min((first / ceiling) * 100, 100))
    second_width = max(1.5, min((second / ceiling) * 100, 100))

    def shown(value: float) -> str:
        if value_kind == "percent":
            return f"{value * 100:.1f}%"
        return f"{value:.1f}".replace(".", ",")

    return f"""
      <div class="comparison-card">
        <strong>{h(ratio)}</strong>
        <div class="comparison-bars">
          <div class="comparison-row first">
            <span>{h(first_label)}</span>
            <div class="comparison-track"><i style="width:{first_width:.2f}%"></i></div>
            <em>{h(shown(first))}</em>
          </div>
          <div class="comparison-row second">
            <span>{h(second_label)}</span>
            <div class="comparison-track"><i style="width:{second_width:.2f}%"></i></div>
            <em>{h(shown(second))}</em>
          </div>
        </div>
        <p>{h(title)}<br><span>{h(reference)}</span></p>
      </div>
    """


def range_card(rows_map: dict[str, dict[str, Any]], name: str, *, wide: bool = False) -> str:
    row = rows_map.get(norm(name), {"name": name})
    fill = bar_fill(row)
    marker = max(0.5, min(fill, 99))
    ref_low, ref_high = parse_pct_reference(row.get("reference_pct") or row.get("relative_reference") or "")
    cls = "range-card wide" if wide else "range-card"
    value = row.get("relative_text") or fmt_pct(row.get("relative"))
    reference = row.get("reference_pct") or row.get("relative_reference") or ""
    return f"""
      <div class="{cls}" style="--value:{marker:.2f}%; --low:{ref_low:.2f}%; --high:{ref_high:.2f}%">
        <div class="range-title">{h(row.get('name') or name)}</div>
        <div class="range-meta"><span>{h(value)}</span><em>VR {h(reference or '-')}</em></div>
        <div class="mini-range"><i></i><b></b></div>
      </div>
    """


def range_section(title: str, cards: list[str], rows_map: dict[str, dict[str, Any]], columns: int = 2) -> str:
    html = "".join(range_card(rows_map, item, wide=(columns == 1)) for item in cards)
    title_html = f'<div class="range-section-title">{h(title)}</div>' if title else ""
    return f"""
      <div class="range-section">
        {title_html}
        <div class="range-grid cols-{columns}">{html}</div>
      </div>
    """


def report_header(patient: dict[str, str], logo: str) -> str:
    return f"""
      <header class="header">
        <img class="logo" src="{asset_url(logo)}" alt="Genoma">
        <div class="address">
          Av. Principal &nbsp; C.C. Guataparo Express, Nivel planta baja, Local 17 y 22,<br>
          Urb. Colinas de Guataparo Nro. 201-130, Valencia Edo. Carabobo<br>
          ZP: 2001 &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Contacto: &nbsp;0424-451.06.32
        </div>
      </header>
      <div class="patient-line">
        <span><b>Fecha</b><u>{h(patient.get('Fecha de muestra', ''))}</u></span>
        <span><b>C.I</b><u>{h(patient.get('Cedula', 'V-'))}</u></span>
        <span><b>Sexo</b><u>{h(patient.get('Sexo', '')).upper()}</u></span>
        <span><b>Nombre</b><u>{h(patient.get('Nombre', ''))}</u></span>
        <span><b>Fecha de nacimiento</b><u>{h(patient.get('Fecha de nacimiento', ''))}</u></span>
      </div>
    """


def pathogen_page(patient: dict[str, str], fondo: str, logo: str) -> str:
    settings = load_pathogen_settings()
    sections = []
    for group in PATHOGEN_GROUPS:
        rows = []
        for key, label in group["items"]:
            status = settings.get(key, "no_detectado")
            rows.append(f"""
              <div class="pathogen-row pathogen-{h(status)}">
                <div class="pathogen-name"><span>{h(label)}</span></div>
                <div class="pathogen-arrow"><i></i></div>
                <div class="pathogen-result">{h(pathogen_status_label(status))}</div>
              </div>
            """)
        note = ""
        if group["title"] == "PATÓGENO GÁSTRICO":
            note = """
              <p class="pathogen-note">*** COMPONENTE GASTRODUODENAL. MANEJO MÉDICO SEGÚN CLÍNICA.</p>
              <p class="pathogen-method">PCR TIEMPO REAL</p>
            """
        sections.append(f"""
          <section class="pathogen-block">
            <h2>{h(group["title"])}</h2>
            <div class="pathogen-list">{''.join(rows)}</div>
            {note}
          </section>
        """)
    return f"""
    <section class="page final pathogen-page">
      <img class="bg" src="{asset_url(fondo)}" alt="">
      {report_header(patient, logo)}
      <div class="green-rule"></div>
      <div class="pathogen-head">
        <h1>Panel de patógenos</h1>
        <p>Resultados cualitativos del componente entérico y gastroduodenal.</p>
      </div>
      <div class="pathogen-layout">{''.join(sections)}</div>
    </section>
    """


def build_extra_pages(data: ExtractedReport, fondo: str, logo: str) -> str:
    rows = data.page2_rows
    rows_map = row_by_name(rows)
    act = pct_text(phylum_total(rows, "1.Actinobacteria"))
    firm = pct_text(phylum_total(rows, "2.Firmicutes"))
    bact = pct_text(phylum_total(rows, "3.Bacteroidetes"))
    prot = pct_text(sum(row_value(rows, n, "relative") or 0 for n in ("Desulfovibrio spp", "E.coli", "Enterobacterales", "Pseudomonas spp")))
    protect = pct_text(sum(row_value(rows, n, "relative") or 0 for n in ("Bifidobacterium spp", "Faecalibacterium prausnitzii", "Akkermansia muciniphila", "Lactobacillaceae")))
    opportun = pct_text(category_total(rows, "5.OPPORTUNISTIC PATHOGENS"))
    candida = pct_text(row_value(rows, "Candida spp", "relative"))

    page4 = detail_table(rows, apply_table_header_settings([
        {"title": "ACTINOBACTERIA", "patient": act, "reference": "30-60%", "rows": [
            "Bifidobacterium adolescentis", "Bifidobacterium animalis subsp. lactis", "Bifidobacterium bifidum",
            "Bifidobacterium breve", "Bifidobacterium catenulatum ssp", "Bifidobacterium dentium",
            "Bifidobacterium longum subsp. infantis", "Bifidobacterium longum subsp. longum",
            "Bifidobacterium spp", "Coriobacteriia", 'Metabolically active "adult" bifidobacteria **',
            'Metabolically active "infant" bifidobacteria **', "Metabolically active bifidobacteria species, proportion",
        ]},
        {"title": "FIRMICUTES", "patient": firm, "reference": "20-40%", "rows": [
            "Clostridium difficile gr", "Clostridium leptum gr", "Dialister+Allisonella+Megasphaera+Veillonella",
            "Enterococcus spp", "Erysipelotrichaceae", "Faecalibacterium prausnitzii", "Lachnospiraceae",
            "Lactobacillaceae", "Lactococcus lactis", "Peptoniphilaceae", "Streptococcus spp",
        ]},
        {"title": "BACTEROIDETES", "patient": bact, "reference": "30-50%", "rows": [
            "Prevotella spp", "Parabacteroides spp", "Butyricimonas spp", "Bacteroides spp", "Alistipes spp",
        ]},
        {"title": "FUSOBACTERIA", "rows": ["Fusobacteriaceae"]},
        {"title": "VERRUMICROBIA", "rows": ["Akkermansia muciniphila"]},
    ]), summary=True, catalog=data.catalog)

    page5 = detail_table(rows, apply_table_header_settings([
        {"title": "PROTEOBACTERIAS", "patient": prot, "reference": "<1%", "rows": ["Desulfovibrio spp", "E.coli", "Enterobacterales", "Pseudomonas spp"]},
        {"title": "MICROBIOTA PROTECTORA", "patient": protect, "reference": "30-60%", "rows": [
            "Bifidobacterium adolescentis", "Bifidobacterium breve", "Bifidobacterium catenulatum ssp",
            "Bifidobacterium dentium", "Bifidobacterium longum subsp. infantis", "Faecalibacterium prausnitzii",
            "Akkermansia muciniphila", "Lactobacillaceae", "Bifidobacterium spp", "Clostridium leptum gr", "Lachnospiraceae",
        ]},
        {"title": "BACTERIAS INMUNOESTIMULANTES (BENEFICIOSAS)", "patient": protect, "reference": "40-70%", "rows": ["Lactobacillaceae", "Akkermansia muciniphila", "Bifidobacterium spp", "Faecalibacterium prausnitzii", "Clostridium leptum gr"]},
        {"title": "B. PRODUCTORAS/FACILITADORA DE MUCINA", "patient": "2,4%", "reference": "1-4% / 40-70%", "rows": ["Akkermansia muciniphila", "Bacteroides spp", "Faecalibacterium prausnitzii", "Lachnospiraceae", "Lactobacillaceae", "Parabacteroides spp"]},
        {"title": "BACTERIAS FORMADORAS DE BUTIRATO", "patient": "68,3%", "reference": "40-60%", "rows": ["Lactobacillaceae", "Butyricimonas spp", "Faecalibacterium prausnitzii", "Clostridium leptum gr", "Lachnospiraceae"]},
        {"title": "PATOGENOS OPORTUNISTAS", "patient": opportun, "reference": "", "rows": ["Pseudomonas spp", "Staphylococcus spp", "Clostridium difficile gr", "Clostridium perfringens gr", "Desulfovibrio spp", "Enterococcus spp", "Fusobacteriaceae", "Peptoniphilaceae", "E.coli", "Enterobacterales", "Erysipelotrichaceae"]},
    ]), catalog=data.catalog)

    page6 = detail_table(rows, apply_table_header_settings([
        {"title": "M. PROINFLAMATORIA (POTENCIALMENTE PERJUDICIAL)", "patient": "21,3%", "reference": "<20%", "rows": ["Bacteroides spp", "Clostridium perfringens gr", "Desulfovibrio spp", "E.coli", "Enterobacterales", "Enterococcus spp", "Pseudomonas spp", "Staphylococcus aureus", "Staphylococcus spp"]},
        {"title": "M.PRODUCTORA DE H2S", "patient": "0,7%", "reference": "0-1%", "rows": ["Clostridium perfringens gr", "Desulfovibrio spp", "Enterobacterales", "Erysipelotrichaceae"]},
        {"title": "M. PRODUCTORA DE HISTAMINA", "patient": "0,4%", "reference": "<1%", "rows": ["Clostridium difficile gr", "Clostridium perfringens gr", "E.coli", "Enterobacterales", "Enterococcus spp", "Lactobacillaceae", "Pseudomonas spp", "Staphylococcus aureus", "Staphylococcus spp"]},
        {"title": "M. PRODUCTORA DE TMA Y TMAO", "patient": "1,0%", "reference": "<2%", "rows": ["Pseudomonas spp", "Clostridium perfringens gr", "Staphylococcus spp", "E.coli", "Enterobacterales", "Desulfovibrio spp", "Bacteroides spp", "Clostridium leptum gr", "Alistipes spp"]},
        {"title": "M. PRODUCTORES DE AMONIACO", "patient": "85,5%", "reference": "<70%", "rows": ["Pseudomonas spp", "Staphylococcus spp", "Clostridium perfringens gr", "Desulfovibrio spp", "Enterococcus spp", "E.coli", "Alistipes spp", "Enterobacterales"]},
    ]), catalog=data.catalog)

    page7 = detail_table(rows, apply_table_header_settings([
        {"title": "M. PRODUCTORA DE FENOLES", "patient": "85,5%", "reference": "<60%", "rows": ["Alistipes spp", "Bacteroides spp", "Clostridium perfringens gr", "Desulfovibrio spp", "E.coli", "Enterobacterales", "Enterococcus spp", "Fusobacteriaceae", "Peptoniphilaceae", "Pseudomonas spp", "Staphylococcus spp"]},
        {"title": "M.PRODUCTORES DE ACIDOS BILIARES SECUNDARIOS", "patient": "85,5%", "reference": "<75%", "rows": ["Alistipes spp", "Bacteroides spp", "Clostridium leptum gr", "Desulfovibrio spp", "Lachnospiraceae"]},
        {"title": "M.PRODUCTORA DE SULFATO DE INDOXILO", "patient": "85,7%", "reference": "<60%", "rows": ["Alistipes spp", "Bacteroides spp", "Clostridium perfringens gr", "Desulfovibrio spp", "E.coli", "Enterobacterales", "Peptoniphilaceae", "Pseudomonas spp"]},
        {"title": "MARCADORES DE PATOGENICIDAD Y RESISTENCIA", "patient": "NO HAY SOBRECRECIMIENTO NI RESISTENCIA", "rows": ["Clostridioides difficile", "mecA", "tcdA tcdB", "srr2", "Streptococcus agalactiae", "Staphylococcus aureus", "Pseudomonas spp", "Enterococcus spp"]},
        {"title": "HONGOS Y LEVADURAS", "rows": ["C.albicans", "Candida spp"]},
        {"title": "ARQUEAS", "rows": ["Methanobrevibacter spp"]},
    ]), catalog=data.catalog)

    yeast_records = load_yeast_profile()
    yeast_rows = "".join(
        f"<div class='yeast-row yeast-{h(record.get('status'))}'>"
        f"<span>{h(record.get('name'))}<em>{h(record.get('value') or yeast_status_label(record.get('status', '')))}</em></span>"
        f"<div class='yeast-scale' style='--value:{yeast_marker(record):.2f}%'><i></i></div></div>"
        for record in yeast_records
    )
    page8 = f"""
      <div class="range-logo"><img src="{asset_url(logo)}" alt=""></div>
      <div class="yeast-legend"><b>MICOBIOMA</b><span class="box weak"></span>Debil<span class="box ok"></span>Moderada<span class="box high"></span>Elevado</div>
      <div class="yeast-list">{yeast_rows}</div>
      <div class="range-legend"><span class="box weak"></span>Bajo<span class="box ok"></span>Optimo<span class="box high"></span>Elevado</div>
      {range_section("ACTINOBACTERIAS", ["Bifidobacterium spp", "Metabolically active bifidobacteria species, proportion", 'Metabolically active "infant" bifidobacteria **', "Bifidobacterium longum subsp. infantis"], rows_map, 2)}
    """

    page9 = (
        '<div class="range-page-spacer"></div>'
        + range_section("", ["Bifidobacterium longum subsp. longum", "Bifidobacterium bifidum", "Bifidobacterium breve", 'Metabolically active "adult" bifidobacteria **', "Bifidobacterium adolescentis", "Bifidobacterium catenulatum ssp", "Bifidobacterium animalis subsp. lactis", "Bifidobacterium dentium"], rows_map, 2)
        + range_section("", ["Coriobacteriia"], rows_map, 1)
        + range_section("FIRMICUTES", ["Clostridium leptum gr", "Dialister+Allisonella+Megasphaera+Veillonella", "Faecalibacterium prausnitzii", "Lachnospiraceae", "Lactobacillaceae", "Lactococcus lactis", "Streptococcus spp"], rows_map, 2)
        + range_section("BACTEROIDETES", ["Alistipes spp", "Bacteroides spp", "Butyricimonas spp", "Parabacteroides spp", "Prevotella spp"], rows_map, 2)
    )

    page10 = (
        range_section("OTHER BACTERIA", ["Akkermansia muciniphila", "Desulfovibrio spp", "Methanobrevibacter spp"], rows_map, 2)
        + range_section("OPPORTUNISTIC PATHOGENS", ["Enterococcus spp", "Erysipelotrichaceae", "Clostridium difficile gr", "Clostridium perfringens gr", "Enterobacterales", "E.coli", "Fusobacteriaceae", "Peptoniphilaceae", "Pseudomonas spp", "Staphylococcus spp"], rows_map, 3)
        + range_section("MARKERS OF PATHOGENICITY AND RESISTANCE", ["Clostridioides difficile", "Staphylococcus aureus", "Streptococcus agalactiae"], rows_map, 3)
        + range_section("YEAST FUNGI", ["Candida spp", "C.albicans"], rows_map, 2)
    )

    pages = [
        ("detail-page p4", page4),
        ("detail-page p5", page5),
        ("detail-page p6", page6),
        ("detail-page p7", page7),
        ("range-page p8", page8),
        ("range-page p9", page9),
        ("range-page p10", page10),
    ]
    return "\n".join(f'<section class="page final {cls}"><img class="bg" src="{asset_url(fondo)}" alt="">{html}</section>' for cls, html in pages)


def build_html(
    pdf: Path,
    xlsx: Path,
    data: ExtractedReport,
    metrics: dict[str, str],
    bullets: list[dict[str, Any]],
    logo: str,
    firma: str,
    fondo: str,
    patient_overrides: dict[str, str] | None = None,
) -> str:
    page1_rows = "\n".join(
        f"<tr class=\"{h(r['kind'])}\"><td>{h(r['name'])}</td><td>{h(r['result'])}</td><td>{h(r['reference'])}</td><td>{h(r['unit'])}</td></tr>"
        for r in final_page1_rows(data.page1_rows)
    )
    bullet_rows = []
    for row in bullets:
        pct = (row.get("relative") or 0) * 100 if row.get("relative") is not None else (row.get("value") or 0) * 10
        pct = max(0, min(float(pct or 0), 100))
        value = fmt_pct(row.get("relative")) if row.get("relative") is not None else h(row.get("value") if row.get("value") is not None else "-")
        bullet_rows.append(
            f"""
            <div class="bullet-row">
              <div class="bullet-name">{h(row['name'])}</div>
              <div class="bullet-track"><span style="width: {pct:.1f}%"></span></div>
              <div class="bullet-value">{value}</div>
            </div>
            """
        )
    mini_rows = []
    for row in data.page2_rows[:18]:
        mini_rows.append(
            f"<tr><td>{h(row['name'])}</td><td>{h(row.get('value') if row.get('value') is not None else '-')}</td>"
            f"<td>{fmt_pct(row.get('relative'))}</td><td>{h(row.get('category'))}</td></tr>"
        )
    patient = dict(data.patient)
    patient.update(patient_overrides or {})
    pathogen_html = pathogen_page(patient, fondo, logo)
    extra_pages = build_extra_pages(data, fondo, logo)
    considerations_html = free_text_to_html(load_considerations())
    clinical_settings = load_clinical_settings()
    stool_macro_html = lines_to_html(clinical_settings.get("stool_macro"))
    stool_micro_html = lines_to_html(clinical_settings.get("stool_micro"))
    firm_value = phylum_total(data.page2_rows, "2.Firmicutes") or named_total(data.page2_rows, FIRMICUTES_NAMES)
    bact_value = phylum_total(data.page2_rows, "3.Bacteroidetes") or named_total(data.page2_rows, BACTEROIDETES_NAMES)
    bifido_value = row_value(data.page2_rows, "Bifidobacterium spp", "value")
    entero_value = row_value(data.page2_rows, "Enterobacterales", "value")
    prev_value = row_value(data.page2_rows, "Prevotella spp", "relative")
    bacteroides_value = row_value(data.page2_rows, "Bacteroides spp", "relative")

    context_query = urlencode({
        "pdf": str(pdf),
        "xlsx": str(xlsx),
        "patient_nombre": patient.get("Nombre", ""),
        "patient_fecha": patient.get("Fecha de muestra", ""),
        "patient_ci": patient.get("Cedula", ""),
        "patient_sexo": patient.get("Sexo", ""),
        "patient_nacimiento": patient.get("Fecha de nacimiento", ""),
    })
    return f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Genoma - reporte preliminar</title>
  <link rel="stylesheet" href="/static/styles.css">
</head>
<body>
  <div class="toolbar" id="appToolbar">
    <input class="toolbar-state" id="toolbarState" type="checkbox" aria-hidden="true">
    <div class="toolbar-compact">
      <div class="toolbar-identity">
        <strong>Genoma PDF Studio</strong>
        <span>{h(pdf.name if str(pdf) else 'Selecciona un PDF')}</span>
      </div>
      <div class="toolbar-quick-actions">
        <button type="button" class="primary-action" onclick="downloadPdf()">Generar PDF</button>
        <label class="ghost-toggle" for="toolbarState" aria-controls="toolbarPanel">
          <span class="show-controls">Mostrar controles</span>
          <span class="hide-controls">Ocultar controles</span>
        </label>
      </div>
    </div>
    <div class="toolbar-panel" id="toolbarPanel">
    <div class="brand-strip">
      <div>
        <strong>Genoma PDF Studio</strong>
        <span>Extrae, valida y genera el informe completo</span>
      </div>
      <nav>
        <a href="/validate?pdf={h(pdf)}&xlsx={h(xlsx)}">Validar extracción</a>
        <a href="/catalog?{h(context_query)}">Valores del sistema</a>
        <a href="/micobioma?{h(context_query)}">Micobioma</a>
        <a href="/clinica?{h(context_query)}">Datos clínicos</a>
        <a href="/consideraciones?{h(context_query)}">Consideraciones</a>
      </nav>
    </div>
    <form method="get">
      <input id="pdfPath" name="pdf" type="hidden" value="{h(pdf)}">
      <input id="xlsxPath" name="xlsx" type="hidden" value="{h(xlsx)}">
      <div class="workflow-card file-status">
        <b>Documento cargado</b>
        <span>{h(pdf.name if str(pdf) else 'Selecciona un PDF')}</span>
      </div>
      <div class="toolbar-actions">
        <button type="submit">Actualizar vista</button>
        <button type="button" class="primary-action" onclick="downloadPdf()">Generar PDF</button>
      </div>
      <div class="patient-editor">
        <label>Nombre <input name="patient_nombre" data-patient-field value="{h(patient.get('Nombre', ''))}"></label>
        <label>Fecha <input name="patient_fecha" data-patient-field value="{h(patient.get('Fecha de muestra', ''))}"></label>
        <label>C.I <input name="patient_ci" data-patient-field value="{h(patient.get('Cedula', 'V-'))}"></label>
        <label>Sexo <input name="patient_sexo" data-patient-field value="{h(patient.get('Sexo', '')).upper()}"></label>
        <label>Fecha de nacimiento <input name="patient_nacimiento" data-patient-field value="{h(patient.get('Fecha de nacimiento', ''))}"></label>
      </div>
    </form>
    <form method="post" action="/upload" enctype="multipart/form-data">
      <label class="upload-drop">Cargar nuevo PDF <input name="pdf_file" type="file" accept="application/pdf"></label>
      <input name="xlsx_mode" type="hidden" value="{MANUAL_SOURCE}">
      <button type="submit">Usar este documento</button>
    </form>
    </div>
  </div>
  <script>
    function downloadPdf() {{
      const params = new URLSearchParams({{
        pdf: document.getElementById('pdfPath').value,
        xlsx: document.getElementById('xlsxPath').value
      }});
      document.querySelectorAll('[data-patient-field]').forEach((input) => {{
        params.set(input.name, input.value);
      }});
      window.location.href = '/export.pdf?' + params.toString();
    }}
  </script>

  <main class="report">
    <section class="page final home">
      <img class="bg" src="{asset_url(fondo)}" alt="">
      <header class="header">
        <img class="logo" src="{asset_url(logo)}" alt="Genoma">
        <div class="address">
          Av. Principal &nbsp; C.C. Guataparo Express, Nivel planta baja, Local 17 y 22,<br>
          Urb. Colinas de Guataparo Nro. 201-130, Valencia Edo. Carabobo<br>
          ZP: 2001 &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Contacto: &nbsp;0424-451.06.32
        </div>
      </header>

      <div class="patient-line">
        <span><b>Fecha</b><u>{h(patient.get('Fecha de muestra', ''))}</u></span>
        <span><b>C.I</b><u>{h(patient.get('Cedula', 'V-'))}</u></span>
        <span><b>Sexo</b><u>{h(patient.get('Sexo', '')).upper()}</u></span>
        <span><b>Nombre</b><u>{h(patient.get('Nombre', ''))}</u></span>
        <span><b>Fecha de nacimiento</b><u>{h(patient.get('Fecha de nacimiento', ''))}</u></span>
      </div>

      <div class="home-warning">IMPORTANT: the interpretation of the results must be made strictly by the attending physician</div>
      <table class="main-table">
        <thead>
          <tr>
            <th>NAME OF RESEARCH</th>
            <th>RESULT</th>
            <th>REFERENCE INTERVAL</th>
            <th>UNIT MEASUREMENTS</th>
          </tr>
        </thead>
        <tbody>{page1_rows}</tbody>
      </table>

      <footer class="footer">
        <div class="bio">MSc. Gabriela Espinoza<br><span>• Genética Clínica<br>• Microbiota Humana<br>• Ciencias Alimentarias</span></div>
        <div class="signature-line"><span>The study was carried out by</span><i></i><img src="{asset_url(firma)}" alt=""></div>
        <div class="legal">LAS CONSIDERACIONES REFLEJADAS EN ESTE ESTUDIO SON ORIENTATIVAS Y SE BASAN EN LOS RESULTADOS OBTENIDOS POR ANÁLISIS MOLECULAR. LA INDICACIÓN FINAL DEBE SER EVALUADA POR EL MÉDICO TRATANTE EN FUNCIÓN DEL CONTEXTO CLÍNICO DEL PACIENTE.</div>
        <div class="notes">* Lg X values are given – means 10 X<br>** the value of the proportion of the number of metabolically active bifidobacteria is indicated</div>
        <div class="bottom-name">{h(patient.get('Nombre', '')).split(' ')[0] if patient.get('Nombre') else ''}</div>
      </footer>
    </section>

    {pathogen_html}

    <section class="page final intro">
      <img class="bg" src="{asset_url(fondo)}" alt="">
      <header class="header">
        <img class="logo" src="{asset_url(logo)}" alt="Genoma">
        <div class="address">
          Av. Principal &nbsp; C.C. Guataparo Express, Nivel planta baja, Local 17 y 22,<br>
          Urb. Colinas de Guataparo Nro. 201-130, Valencia Edo. Carabobo<br>
          ZP: 2001 &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Contacto: &nbsp;0424-451.06.32
        </div>
      </header>

      <div class="patient-line">
        <span><b>Fecha</b><u>{h(patient.get('Fecha de muestra', ''))}</u></span>
        <span><b>C.I</b><u>{h(patient.get('Cedula', 'V-'))}</u></span>
        <span><b>Sexo</b><u>{h(patient.get('Sexo', '')).upper()}</u></span>
        <span><b>Nombre</b><u>{h(patient.get('Nombre', ''))}</u></span>
        <span><b>Fecha de nacimiento</b><u>{h(patient.get('Fecha de nacimiento', ''))}</u></span>
      </div>

      <div class="green-rule"></div>
      <div class="considerations">
        <div class="free-considerations">
          {considerations_html}
        </div>
        <div class="legacy-considerations">
        <p><b>Consideraciones:</b> DISBIOSIS METABÓLICA LEVE NO INFLAMATORIA</p>
        <p>1-Biomasa total dentro de los parámetros para la edad.<br>
        2-Excelente capacidad antinflamatoria colónica.<br>
        3-Perfil funcional proteolítico con elevación de metabolitos clave y ácidos biliares secundarios, aún sin daño mucoso.<br>
        4-Diversidad bacteriana adecuada a la edad.<br>
        5-Déficit de simbiontes protectores.<br>
        6-Barrera intestinal protegida.<br>
        7-Excelente potencial colónico funcional.<br>
        8-Sin riesgo de permeabilidad intestinal.<br>
        9-Leve sobre-crecimiento micótico,</p>
        <p><b>Recomendaciones nutricionales generales:</b></p>
        <p>Con el objetivo de modular la microbiota, se recomienda implementar una alimentación antiinflamatoria, basada en vegetales cocidos, frutas bajas en FODMAPs, y proteínas de alto valor biológico, preferiblemente pescado. Debe evitar durante 8 semanas el consumo de carne roja, huevos, embutidos, trigo, vinagre balsámico, refrescos, azucares refinados, lácteos sin no son fermentados, antibióticos sin prescripción medica, edulcorantes y alimentos ultraprocesados e industrializados.</p>
        <p><b>Considerando la edad del paciente, aumentar progresivamente el consumo de:</b><br>
        • Frutos rojos, kombucha sin azúcar agregado, yogurt natural sin azúcar, cacao, aceite de oliva, aceite de coco, orégano, canela, vinagre de manzana. Manzana y pera cocida<br>
        • Verduras cocidas ( zanahoria, batata, auyama). Así como también, alcachofas, kale, repollo morado, espárragos, cebolla cruda (microdosis), ajo, mantequilla clarificada, plátano verde cocido, aguacate y frutos secos activados.<br>
        • Higado de res en microdosis.</p>
        <p><b>Objetivos:</b><br>
        • Reducir bacterias proinflamatorias.<br>
        • Restaurar la integridad de la barrera mucosa inmunológica intestinal.<br>
        • Prevenir sobrecrecimiento de oportunistas y hongos.<br>
        • Restaurar la microbiotica simbiótica.</p>
        <p><b>Probióticos recomendados por fase:</b><br>
        <b>Indicaciones:</b> tomar una capsula diaria en ayuna con bebidas fresca no calientes y lejos de antibióticos.<br>
        <b>Fase 1:</b> 8 semanas. De las cuales 2 semanas media dosis, luego continuar dosis completa.<br>
        Bifidobacterium breve, B. bifidum infantis, B. longum.<br>
        Lactobacillus plantarum, L. rhamnosus. 10 mil millones UFC/día. En polvo.<br>
        <b>Fase 2:</b> 12 semanas<br>
        Probióticos de almidón resistente. Pediátrico.1 sobre/día.<br>
        <b>Fase3:</b> 12 semanas<br>
        Akkermansia muciniphila. 1cápsula/semana.<br>
        <b>Fase de mantenimiento:</b> 6 semana de refuerzo<br>
        Saccharomyces boulardii. 5 mil UFC/DÍA. Evaluar tolerancia.<br>
        Se recomienda evaluar zonulina.<br>
        Evaluación de microbiota en 4 meses.</p>
      </div>
      </div>
    </section>

    <section class="page final pr1">
      <img class="bg" src="{asset_url(fondo)}" alt="">
      <div class="pr-bar">ANÁLISIS GENÉTICO MOLECULAR DEL MICROBIOMA</div>
      <div class="pr-bar">PROPIEDADES DE LAS HECES</div>
      <div class="stool-grid">
        <div><b>Descripción macroscópica</b><br>{stool_macro_html}</div>
        <div><b>Descripción microscópica</b><br>{stool_micro_html}</div>
      </div>

      <div class="legend-bar"><b>DIVERSIDAD TAXONOMICA</b><span class="swatch lime"></span>Muy bajo<span class="swatch yellow"></span>Bajo<span class="swatch green"></span>Optimo</div>
      <div class="taxa-row">
        <span>Diversity, number of taxa:</span>
        <strong>{h((metrics.get('diversity') or '-').split()[0])}</strong>
        <div class="range"><span class="lime"></span><span class="yellow"></span><span class="green"></span><i style="left:60%"></i></div>
      </div>
      <p class="center-copy">La diversidad bacteriana es un buen indicador de la salud intestinal. Una mayor diversidad suele estar asociada con una mejor salud digestiva y un mejor sistema inmunológico. Puede disminuir en respuesta a terapias con antibióticos, infecciones, la edad, las dietas desequilibradas o el tabaquismo.</p>

      <div class="pr-bar">ENTEROTIPO</div>
      <p class="copy">El microbioma intestinal esta dividido en tres grupos bacterianos dominantes, asociados con diferentes perfiles metabólicos y funciones en el organismo. Enterotipo 1 Enterotipo 2 Enterotipo 3. {h(clinical_settings.get('enterotype_text', ''))}</p>
      <div class="enterotype-box">{h(clinical_settings.get('enterotype_number', '1'))}</div>

      <div class="legend-bar"><b>MICROBIOTA BENEFICA</b><span class="swatch lime"></span>Muy bajo<span class="swatch yellow"></span>Bajo<span class="swatch green"></span>Optimo</div>
      <div class="taxa-row normal">
        <span>NORMAL MICROBIOTA</span>
        <strong>{h(metrics.get('normal_microbiota'))}</strong>
        <div class="range normal-range"><span class="lime"></span><span class="yellow"></span><span class="green"></span><i style="left:92%"></i></div>
      </div>
      <p class="copy">La proporcion bacteriana en el tracto intestinal puede variar considerablemente de persona a persona. Este indicativo refleja el equilibrio cuantitativo de una microbiota benefica y una microbiota potencialmente patogena.</p>

      <div class="pr-bar metabolic"><span>EQUILIBRIO METABÓLICO INTESTINAL</span><span>*DISBIOSIS PROTEOLITICA.</span></div>
      <div class="ratio-strip">
        {comparison_card(metrics.get('firm_bact'), 'Firmicutes / Bacteroidetes', 'Reference Value: 1.5 - 2.5', 'Firmicutes', firm_value, 'Bacteroidetes', bact_value, value_kind='percent')}
        {comparison_card(metrics.get('actino_proteo'), 'Actinobacteria / Proteobacteria', 'Reference Value > 1.0', 'Bifidobacterium spp', bifido_value, 'Enterobacterales', entero_value, value_kind='lg', scale=9)}
        {comparison_card(metrics.get('prev_bacteroides'), 'Prevotella / Bacteroides', 'Reference Value: 0.1 - 1.0', 'Prevotella', prev_value, 'Bacteroides', bacteroides_value, value_kind='percent')}
      </div>
    </section>
    {extra_pages}
  </main>
</body>
</html>"""


def build_validation_html(pdf: Path, xlsx: Path) -> str:
    data = extract_report(pdf, xlsx)
    expected = [clean_text(row.get("NAME OF RESEARCH")) for row in data.catalog.get("Resultados", []) if row.get("NAME OF RESEARCH")]
    found_norm = {norm(row["name"]) for row in data.page2_rows}
    missing = [name for name in expected if norm(name) not in found_norm]
    metrics = calculate_metrics(data.page2_rows)

    page1_rows = "\n".join(
        f"<tr><td>{i+1}</td><td>{h(r['name'])}</td><td>{h(r['result'])}</td><td>{h(r['reference'])}</td><td>{h(r['unit'])}</td></tr>"
        for i, r in enumerate(data.page1_rows)
    )
    page2_rows = "\n".join(
        f"<tr><td>{i+1}</td><td>{h(r['name'])}</td><td>{h(r.get('absolute_text'))}</td><td>{h(r.get('absolute_reference'))}</td>"
        f"<td>{h(r.get('relative_text'))}</td><td>{h(r.get('relative_reference'))}</td><td>{h(r.get('category'))}</td><td>{h(r.get('raw'))}</td></tr>"
        for i, r in enumerate(data.page2_rows)
    )
    missing_rows = "".join(f"<li>{h(name)}</li>" for name in missing) or "<li>Ninguno</li>"
    return f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Validacion extraccion PDF</title>
  <link rel="stylesheet" href="/static/styles.css">
</head>
<body>
  <div class="toolbar">
    <form method="get">
      <label>PDF <input name="pdf" value="{h(pdf)}"></label>
      <label>Base de valores <input name="xlsx" value="{h(xlsx)}"></label>
      <button type="submit">Validar</button>
      <a class="link-button" href="/?pdf={h(pdf)}&xlsx={h(xlsx)}">Ver reporte</a>
    </form>
  </div>
  <main class="validation">
    <h1>Validacion de extraccion</h1>
    <section class="validation-card">
      <h2>Resumen</h2>
      <p><b>Paciente:</b> {h(data.patient.get('Nombre'))} | <b>Sexo:</b> {h(data.patient.get('Sexo'))} | <b>Nacimiento:</b> {h(data.patient.get('Fecha de nacimiento'))} | <b>Muestra:</b> {h(data.patient.get('Fecha de muestra'))}</p>
      <p><b>PAGE001:</b> {len(data.page1_rows)} filas | <b>PAGE002:</b> {len(data.page2_rows)} de {len(expected)} nombres esperados | <b>Faltantes:</b> {len(missing)}</p>
      <p><b>Metricas:</b> Firmicutes/Bacteroidetes = {h(metrics.get('firm_bact'))}; Actino/Proteo = {h(metrics.get('actino_proteo'))}; Prevotella/Bacteroides = {h(metrics.get('prev_bacteroides'))}</p>
    </section>
    <section class="validation-card">
      <h2>Observacion extraida</h2>
      <pre>{h(data.observation)}</pre>
    </section>
    <section class="validation-card">
      <h2>Nombres faltantes contra la base interna</h2>
      <ul>{missing_rows}</ul>
    </section>
    <section class="validation-card">
      <h2>PAGE001 - primera pagina del PDF</h2>
      <table class="validation-table"><thead><tr><th>#</th><th>Nombre</th><th>Resultado</th><th>Referencia</th><th>Unidad</th></tr></thead><tbody>{page1_rows}</tbody></table>
    </section>
    <section class="validation-card">
      <h2>PAGE002 - tabla microbioma por coordenadas</h2>
      <table class="validation-table"><thead><tr><th>#</th><th>Nombre</th><th>Abs.</th><th>Ref. abs.</th><th>Rel. %</th><th>Ref. rel.</th><th>Categoria</th><th>Linea cruda</th></tr></thead><tbody>{page2_rows}</tbody></table>
    </section>
  </main>
</body>
</html>"""


def build_catalog_html(saved: bool = False) -> str:
    records = load_manual_catalog().get("Resultados", [])
    rows = list(records)
    rows.extend({} for _ in range(5))
    header_cells = "".join(f"<th>{h(label)}</th>" for _, label in CATALOG_FORM_FIELDS)
    def select_cell(field: str, value: str, options: list[tuple[str, str]]) -> str:
        known = {key for key, _ in options}
        option_html = "".join(
            f"<option value=\"{h(key)}\"{' selected' if value == key else ''}>{h(label)}</option>"
            for key, label in options
        )
        if value and value not in known:
            option_html += f"<option value=\"{h(value)}\" selected>{h(value)}</option>"
        return f"<td><select name=\"{h(field)}\">{option_html}</select></td>"

    def report_tables_cell(value: str = "") -> str:
        option_html = '<option value="">Agregar ficha...</option>' + "".join(
            f"<option value=\"{h(option)}\">{h(option)}</option>"
            for option in REPORT_TABLE_OPTIONS
        )
        return f"""
          <td>
            <div class="table-chip-picker" data-table-picker>
              <input type="hidden" name="report_tables" value="{h(value)}">
              <div class="table-chip-list" data-chip-list></div>
              <div class="table-chip-add">
                <select data-table-select>{option_html}</select>
                <button type="button" class="chip-add-button" data-table-add aria-label="Agregar ficha">+</button>
              </div>
            </div>
          </td>
        """

    rows_html = []
    for index, row in enumerate(rows, start=1):
        inputs = []
        for field, label in CATALOG_FORM_FIELDS:
            value = row.get(field, "")
            if not value and field == "category":
                value = row.get("Category", "")
            elif not value and field == "category_pg1":
                value = row.get("Categoria Asignada Pg1", "")
            elif not value and field == "reference":
                value = row.get("Reference interval", "")
            elif not value and field == "reference_pct":
                value = row.get("Reference interval %", "")
            elif not value and field == "report_tables":
                value = row.get("Report tables", "")
            elif not value and field == "display_order":
                value = row.get("Display order", "")
            if field == "category":
                inputs.append(select_cell(field, value, FUNCTIONAL_CATEGORY_OPTIONS))
            elif field == "category_pg1":
                inputs.append(select_cell(field, value, PHYLA_CATEGORY_OPTIONS))
            elif field == "report_tables":
                inputs.append(report_tables_cell(value))
            elif field == "display_order":
                inputs.append(
                    f"<td><input class=\"order-input\" name=\"{h(field)}\" value=\"{h(value)}\" "
                    f"type=\"number\" step=\"1\" placeholder=\"10\"></td>"
                )
            else:
                inputs.append(
                    f"<td><input name=\"{h(field)}\" value=\"{h(value)}\" "
                    f"placeholder=\"{h(label)}\"></td>"
                )
        rows_html.append(f"<tr><th scope=\"row\">{index}</th>{''.join(inputs)}</tr>")
    rows_markup = "\n".join(rows_html)
    notice = "<div class='notice success'>Valores guardados. Ya puedes generar reportes usando la opcion de valores guardados.</div>" if saved else ""
    template_cells = "".join(
        select_cell(field, "", FUNCTIONAL_CATEGORY_OPTIONS) if field == "category"
        else select_cell(field, "", PHYLA_CATEGORY_OPTIONS) if field == "category_pg1"
        else report_tables_cell("") if field == "report_tables"
        else f'<td><input class="order-input" name="{h(field)}" type="number" step="1" placeholder="10"></td>' if field == "display_order"
        else f'<td><input name="{h(field)}" placeholder="{h(label)}"></td>'
        for field, label in CATALOG_FORM_FIELDS
    )
    return f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Valores del sistema - Genoma</title>
  <link rel="stylesheet" href="/static/styles.css">
</head>
<body class="app-screen">
  <main class="settings-shell catalog-settings-shell">
    <section class="settings-hero">
      <div>
        <p>Base interna</p>
        <h1>Valores del sistema</h1>
        <span>Estos datos alimentan el informe. El grupo controla la seccion del reporte; el filo controla las sumatorias y graficos de resumen.</span>
      </div>
      <a class="link-button" href="/">Volver al reporte</a>
    </section>
    {notice}
    <section class="settings-card">
      <div class="settings-meta">
        <strong>{len(records)} filas guardadas</strong>
        <span>Para uso diario normalmente solo revisas nombre y referencias. Cambia grupos solo si agregas un nuevo marcador.</span>
      </div>
      <div class="catalog-help-grid">
        <div><b>Grupo del informe</b><span>Decide en que seccion funcional aparece: Firmicutes, patogenos oportunistas, marcadores, levaduras, etc.</span></div>
        <div><b>Filo / resumen</b><span>Decide donde suma para la tabla superior y ratios: Actinobacteria, Firmicutes, Bacteroidetes, Proteobacteria, etc.</span></div>
        <div><b>Referencias</b><span>Son los rangos que se imprimen junto a cada resultado extraido del PDF.</span></div>
        <div><b>Fichas/tablas</b><span>Muestra en que tablas ya esta cada valor. Puedes agregar varias fichas con el selector y quitarlas con la X.</span></div>
        <div><b>Orden</b><span>Controla la posicion dentro de cada tabla. Usa numeros espaciados como 10, 20, 30 para poder insertar filas despues.</span></div>
      </div>
      <form class="catalog-form" method="post" action="/catalog">
        <div class="catalog-table-wrap">
          <table class="catalog-form-table">
            <thead>
              <tr><th>#</th>{header_cells}</tr>
            </thead>
            <tbody id="catalogRows">
              {rows_markup}
            </tbody>
          </table>
        </div>
        <div class="settings-actions">
          <button type="submit">Guardar valores</button>
          <button type="button" class="ghost-button" id="addCatalogRow">Agregar fila</button>
          <a class="link-button" href="/?xlsx={MANUAL_SOURCE}">Usar en reporte</a>
        </div>
      </form>
    </section>
  </main>
  <template id="catalogRowTemplate">
    <tr>
      <th scope="row"></th>
      {template_cells}
    </tr>
  </template>
  <script>
    const rows = document.querySelector("#catalogRows");
    const template = document.querySelector("#catalogRowTemplate");
    function splitTables(value) {{
      return value.split(",").map((item) => item.trim()).filter(Boolean);
    }}
    function syncTablePicker(picker) {{
      const input = picker.querySelector("input[name='report_tables']");
      const list = picker.querySelector("[data-chip-list]");
      const values = [...new Set(splitTables(input.value))];
      input.value = values.join(", ");
      list.innerHTML = "";
      if (!values.length) {{
        const empty = document.createElement("span");
        empty.className = "table-chip-empty";
        empty.textContent = "Sin fichas";
        list.appendChild(empty);
        return;
      }}
      values.forEach((value) => {{
        const chip = document.createElement("span");
        chip.className = "table-chip";
        chip.textContent = value;
        const remove = document.createElement("button");
        remove.type = "button";
        remove.textContent = "×";
        remove.setAttribute("aria-label", "Quitar " + value);
        remove.addEventListener("click", () => {{
          input.value = splitTables(input.value).filter((item) => item !== value).join(", ");
          syncTablePicker(picker);
        }});
        chip.appendChild(remove);
        list.appendChild(chip);
      }});
    }}
    function initTablePicker(picker) {{
      if (picker.dataset.ready === "1") return;
      picker.dataset.ready = "1";
      const input = picker.querySelector("input[name='report_tables']");
      const select = picker.querySelector("[data-table-select]");
      const add = picker.querySelector("[data-table-add]");
      add.addEventListener("click", () => {{
        const value = select.value.trim();
        if (!value) return;
        const values = splitTables(input.value);
        if (!values.includes(value)) values.push(value);
        input.value = values.join(", ");
        select.value = "";
        syncTablePicker(picker);
      }});
      select.addEventListener("change", () => add.click());
      syncTablePicker(picker);
    }}
    function renumberCatalogRows() {{
      rows.querySelectorAll("tr").forEach((row, index) => {{
        row.querySelector("th").textContent = index + 1;
      }});
    }}
    document.querySelectorAll("[data-table-picker]").forEach(initTablePicker);
    document.querySelector("#addCatalogRow").addEventListener("click", () => {{
      const newRow = template.content.firstElementChild.cloneNode(true);
      rows.appendChild(newRow);
      newRow.querySelectorAll("[data-table-picker]").forEach(initTablePicker);
      renumberCatalogRows();
      rows.lastElementChild.querySelector("input").focus();
    }});
  </script>
</body>
</html>"""
    notice = "<div class='notice success'>Valores guardados. Ya puedes generar reportes usando la opción de valores guardados.</div>" if saved else ""
    return f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Valores del sistema - Genoma</title>
  <link rel="stylesheet" href="/static/styles.css">
</head>
<body class="app-screen">
  <main class="settings-shell">
    <section class="settings-hero">
      <div>
        <p>Base interna</p>
        <h1>Valores del sistema</h1>
        <span>Estos datos reemplazan al Excel cuando eliges “Usar valores guardados”. Pega aquí la tabla desde Excel en formato tabulado.</span>
      </div>
      <a class="link-button" href="/">Volver al reporte</a>
    </section>
    {notice}
    <section class="settings-card">
      <div class="settings-meta">
        <strong>{len(records)} filas guardadas</strong>
        <span>Columnas esperadas: NAME OF RESEARCH, category, category_pg1, reference, reference_pct</span>
      </div>
      <form method="post" action="/catalog">
        <textarea name="catalog_text" spellcheck="false">{h(text)}</textarea>
        <div class="settings-actions">
          <button type="submit">Guardar valores</button>
          <a class="link-button" href="/?xlsx={MANUAL_SOURCE}">Usar en reporte</a>
        </div>
      </form>
    </section>
    <section class="settings-card compact">
      <h2>Ejemplo de formato</h2>
      <pre>{h(sample)}</pre>
    </section>
  </main>
</body>
</html>"""


def build_yeast_html(saved: bool = False) -> str:
    records = load_yeast_profile()
    rows_html = []
    for index, record in enumerate(records, start=1):
        options = "".join(
            f"<option value=\"{h(key)}\"{' selected' if record.get('status') == key else ''}>{h(label)}</option>"
            for key, label in YEAST_STATUSES
        )
        rows_html.append(f"""
          <tr>
            <th scope="row">{index}</th>
            <td>
              <input type="hidden" name="name" value="{h(record.get('name'))}">
              <strong>{h(record.get('name'))}</strong>
            </td>
            <td><select name="status">{options}</select></td>
            <td><input name="value" value="{h(record.get('value'))}" placeholder="0-100"></td>
            <td><input name="notes" value="{h(record.get('notes'))}" placeholder="Opcional"></td>
          </tr>
        """)
    notice = "<div class='notice success'>Micobioma guardado. El proximo PDF usara estos marcadores.</div>" if saved else ""
    return f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Micobioma - Genoma</title>
  <link rel="stylesheet" href="/static/styles.css">
</head>
<body class="app-screen">
  <main class="settings-shell catalog-settings-shell">
    <section class="settings-hero">
      <div>
        <p>Valores graficados</p>
        <h1>Micobioma</h1>
        <span>Estos campos controlan las barras de Candida, Malassezia, Pichia, Saccharomyces y otros hongos de la pagina de micobioma.</span>
      </div>
      <a class="link-button" href="/">Volver al reporte</a>
    </section>
    {notice}
    <section class="settings-card">
      <div class="settings-meta">
        <strong>{len(records)} marcadores</strong>
        <span>Usa estado clinico o coloca un valor 0-100 para posicionar exactamente el marcador en la barra.</span>
      </div>
      <form method="post" action="/micobioma">
        <div class="catalog-table-wrap">
          <table class="catalog-form-table yeast-form-table">
            <thead><tr><th>#</th><th>Hongo / levadura</th><th>Estado</th><th>Valor grafico</th><th>Nota</th></tr></thead>
            <tbody>{''.join(rows_html)}</tbody>
          </table>
        </div>
        <div class="settings-actions">
          <button type="submit">Guardar micobioma</button>
          <a class="link-button" href="/">Usar en reporte</a>
        </div>
      </form>
    </section>
  </main>
</body>
</html>"""


def build_considerations_html(saved: bool = False) -> str:
    text = load_considerations()
    notice = "<div class='notice success'>Consideraciones guardadas. El proximo PDF usara este texto.</div>" if saved else ""
    return f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Consideraciones - Genoma</title>
  <link rel="stylesheet" href="/static/styles.css">
</head>
<body class="app-screen">
  <main class="settings-shell">
    <section class="settings-hero considerations-hero">
      <div>
        <p>Texto libre</p>
        <h1>Consideraciones</h1>
        <span>La doctora escribe aqui las consideraciones, recomendaciones y objetivos antes de generar el PDF. Se respetan saltos de linea y parrafos.</span>
      </div>
      <a class="link-button" href="/">Volver al reporte</a>
    </section>
    {notice}
    <section class="settings-card">
      <div class="settings-meta">
        <strong>Pagina 2 del informe</strong>
        <span>Deja una linea en blanco para separar parrafos. Las lineas que terminan en dos puntos se muestran como encabezados.</span>
      </div>
      <form method="post" action="/consideraciones">
        <textarea class="clinical-notes" name="considerations_text" spellcheck="true">{h(text)}</textarea>
        <div class="settings-actions">
          <button type="submit">Guardar consideraciones</button>
          <a class="link-button" href="/">Usar en reporte</a>
        </div>
      </form>
    </section>
  </main>
</body>
</html>"""


def build_clinical_html(saved: bool = False) -> str:
    settings = load_clinical_settings()
    pathogen_settings = load_pathogen_settings()
    table_header_settings = load_table_header_settings()
    pathogen_sections = []
    for group in PATHOGEN_GROUPS:
        rows = []
        for key, label in group["items"]:
            selected_value = pathogen_settings.get(key, "no_detectado")
            options = "\n".join(
                f'<option value="{h(value)}"{" selected" if selected_value == value else ""}>{h(text)}</option>'
                for value, text in PATHOGEN_STATUS_OPTIONS
            )
            rows.append(f"""
              <label class="pathogen-setting-row">
                <span>{h(label)}</span>
                <select name="pathogen_{h(key)}">{options}</select>
              </label>
            """)
        pathogen_sections.append(f"""
          <div class="pathogen-settings-block">
            <h3>{h(group["title"])}</h3>
            <div class="pathogen-settings-grid">{''.join(rows)}</div>
          </div>
        """)
    table_header_rows = []
    for title, default_patient, default_reference in TABLE_HEADER_SPECS:
        key = table_header_key(title)
        values = table_header_settings.get(key, {"patient": "", "reference": ""})
        table_header_rows.append(f"""
          <div class="table-header-setting-row">
            <div>
              <strong>{h(title)}</strong>
              <span>Si queda vacío, usa el valor automático o el valor base: Paciente {h(default_patient or 'automático')} · V.R {h(default_reference or 'sin V.R')}</span>
            </div>
            <label>Paciente
              <input name="table_patient_{h(key)}" value="{h(values.get('patient', ''))}" placeholder="{h(default_patient or 'automático')}">
            </label>
            <label>V.R
              <input name="table_reference_{h(key)}" value="{h(values.get('reference', ''))}" placeholder="{h(default_reference or 'sin V.R')}">
            </label>
          </div>
        """)
    notice = "<div class='notice success'>Datos clínicos guardados. El reporte usará estos valores.</div>" if saved else ""
    return f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Datos clínicos - Genoma</title>
  <link rel="stylesheet" href="/static/styles.css">
</head>
<body class="app-screen">
  <main class="settings-shell">
    <section class="settings-hero">
      <div>
        <p>Configuración clínica</p>
        <h1>Datos clínicos</h1>
        <span>Estos campos alimentan textos interpretativos, propiedades de heces y el panel dinámico de patógenos.</span>
      </div>
      <a class="link-button" href="/">Volver al reporte</a>
    </section>
    {notice}
    <section class="settings-card">
      <div class="settings-meta">
        <strong>Enterotipo del informe</strong>
        <span>El número aparece dentro del recuadro visual y el texto se imprime en la sección ENTEROTIPO.</span>
      </div>
      <form method="post" action="/clinica">
        <div class="clinical-grid">
          <label>Enterotipo
            <select name="enterotype_number">
              <option value="1"{' selected' if settings.get('enterotype_number') == '1' else ''}>1</option>
              <option value="2"{' selected' if settings.get('enterotype_number') == '2' else ''}>2</option>
              <option value="3"{' selected' if settings.get('enterotype_number') == '3' else ''}>3</option>
            </select>
          </label>
          <label>Descripción corta
            <input name="enterotype_name" value="{h(settings.get('enterotype_name', ''))}" placeholder="Proteolítico">
          </label>
        </div>
        <label class="clinical-textarea-label">Texto que aparecerá en el informe
          <textarea name="enterotype_text" spellcheck="true">{h(settings.get('enterotype_text', ''))}</textarea>
        </label>
        <div class="settings-meta clinical-subtitle">
          <strong>Propiedades de las heces</strong>
          <span>Estos textos se imprimen en la página de análisis genético molecular.</span>
        </div>
        <div class="clinical-grid stool-settings">
          <label>Descripción macroscópica
            <textarea name="stool_macro" spellcheck="true">{h(settings.get('stool_macro', ''))}</textarea>
          </label>
          <label>Descripción microscópica
            <textarea name="stool_micro" spellcheck="true">{h(settings.get('stool_micro', ''))}</textarea>
          </label>
        </div>
        <div class="settings-meta clinical-subtitle">
          <strong>Panel de patógenos</strong>
          <span>Estos resultados construyen la nueva segunda página del PDF.</span>
        </div>
        <div class="pathogen-settings">
          {''.join(pathogen_sections)}
        </div>
        <div class="settings-meta clinical-subtitle">
          <strong>Encabezados de tablas</strong>
          <span>Edita lo que aparece como *PACIENTE y *V.R en los títulos de las tablas. Puedes escribir porcentajes, rangos o texto libre.</span>
        </div>
        <div class="table-header-settings">
          {''.join(table_header_rows)}
        </div>
        <div class="settings-actions">
          <button type="submit">Guardar datos clínicos</button>
          <a class="link-button" href="/">Usar en reporte</a>
        </div>
      </form>
    </section>
  </main>
</body>
</html>"""


def build_upload_required_html(pdf: Path, xlsx: Path) -> str:
    return f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Genoma PDF Studio</title>
  <link rel="stylesheet" href="/static/styles.css">
</head>
<body class="app-screen">
  <main class="settings-shell">
    <section class="settings-hero">
      <div>
        <p>Genoma PDF Studio</p>
        <h1>Cargar informe de microbioma</h1>
        <span>Sube el PDF del laboratorio para extraer los datos y generar el informe completo.</span>
      </div>
      <a class="link-button" href="/catalog">Valores del sistema</a>
    </section>
    <section class="settings-card">
      <div class="settings-meta">
        <strong>No hay PDF cargado</strong>
        <span>La app esta lista. El primer paso es cargar un PDF del laboratorio.</span>
      </div>
      <form method="post" action="/upload" enctype="multipart/form-data">
        <label class="upload-drop">Cargar PDF <input name="pdf_file" type="file" accept="application/pdf" required></label>
        <input name="xlsx_mode" type="hidden" value="{MANUAL_SOURCE}">
        <div class="settings-actions">
          <button type="submit">Usar este documento</button>
          <a class="link-button" href="/consideraciones">Consideraciones</a>
          <a class="link-button" href="/micobioma">Micobioma</a>
          <a class="link-button" href="/clinica">Datos clínicos</a>
        </div>
      </form>
    </section>
  </main>
</body>
</html>"""


def chrome_path() -> Path:
    for candidate in CHROME_CANDIDATES:
        if candidate.exists():
            return candidate
    found = shutil.which("chrome") or shutil.which("msedge") or shutil.which("chromium") or shutil.which("chromium-browser")
    if found:
        return Path(found)
    raise RuntimeError("No encontre Chrome o Edge para generar el PDF.")


def report_url(pdf: Path, xlsx: Path, patient_overrides: dict[str, str] | None = None) -> str:
    patient_params = {
        "Nombre": "patient_nombre",
        "Fecha de muestra": "patient_fecha",
        "Cedula": "patient_ci",
        "Sexo": "patient_sexo",
        "Fecha de nacimiento": "patient_nacimiento",
    }
    payload = {"pdf": str(pdf), "xlsx": str(xlsx), "print": "1"}
    for field, param in patient_params.items():
        if patient_overrides and field in patient_overrides:
            payload[param] = patient_overrides[field]
    query = urlencode(payload)
    internal_port = int(os.environ.get("PORT", "5055"))
    return f"http://127.0.0.1:{internal_port}/?{query}"


def render_pdf_with_chrome(pdf: Path, xlsx: Path, patient_overrides: dict[str, str] | None = None) -> bytes:
    with tempfile.TemporaryDirectory(prefix="genoma_pdf_") as tmp:
        out = Path(tmp) / "genoma_reporte.pdf"
        profile = Path(tmp) / "chrome-profile"
        cmd = [
            str(chrome_path()),
            "--headless=new",
            "--disable-gpu",
            "--disable-dev-shm-usage",
            "--no-sandbox",
            "--no-pdf-header-footer",
            f"--user-data-dir={profile}",
            f"--print-to-pdf={out}",
            report_url(pdf, xlsx, patient_overrides),
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
        if result.returncode != 0 or not out.exists():
            raise RuntimeError(result.stderr.strip() or "Chrome no pudo crear el PDF.")
        from pypdf import PdfReader, PdfWriter

        reader = PdfReader(str(out))
        writer = PdfWriter()
        for page in reader.pages:
            page.scale_to(648, 828)
            writer.add_page(page)
        buffer = io.BytesIO()
        writer.write(buffer)
        return buffer.getvalue()


def safe_upload_name(filename: str) -> str:
    name = Path(filename or "archivo").name
    name = re.sub(r"[^\w.\- ()\[\]ñÑáéíóúÁÉÍÓÚ]+", "_", name, flags=re.U).strip(" ._")
    return name or "archivo"


class Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        try:
            if path == "/health":
                body = b"ok"
                self.send_response(200)
                self.send_header("Content-Type", "text/plain; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
                return
            if path.startswith("/assets/"):
                self.send_file(RESOURCES / path.removeprefix("/assets/"))
                return
            if path.startswith("/static/"):
                self.send_file(APP_DIR / path.removeprefix("/"))
                return
            params = parse_qs(parsed.query)
            pdf = Path(params.get("pdf", [str(default_pdf_path())])[0])
            xlsx = Path(params.get("xlsx", [str(default_xlsx_path())])[0])
            patient_overrides = merged_patient_overrides(params)
            if "pdf" in params or "xlsx" in params or patient_overrides_from_params(params):
                save_session_state(pdf=pdf, xlsx=xlsx, patient_overrides=patient_overrides)
            if path == "/catalog":
                html = build_catalog_html(saved=params.get("saved", ["0"])[0] == "1").encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(html)))
                self.end_headers()
                self.wfile.write(html)
                return
            if path == "/micobioma":
                html = build_yeast_html(saved=params.get("saved", ["0"])[0] == "1").encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(html)))
                self.end_headers()
                self.wfile.write(html)
                return
            if path == "/clinica":
                html = build_clinical_html(saved=params.get("saved", ["0"])[0] == "1").encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(html)))
                self.end_headers()
                self.wfile.write(html)
                return
            if path == "/consideraciones":
                html = build_considerations_html(saved=params.get("saved", ["0"])[0] == "1").encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(html)))
                self.end_headers()
                self.wfile.write(html)
                return
            if path == "/export.pdf":
                data = render_pdf_with_chrome(pdf, xlsx, patient_overrides)
                filename = safe_upload_name(f"{pdf.stem or 'genoma'}_GENOMA.pdf")
                self.send_response(200)
                self.send_header("Content-Type", "application/pdf")
                self.send_header("Content-Disposition", f'attachment; filename="{quote(filename)}"')
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return
            if path == "/validate":
                html = build_validation_html(pdf, xlsx).encode("utf-8")
            else:
                html = render_report(pdf, xlsx, patient_overrides).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(html)))
            self.end_headers()
            self.wfile.write(html)
        except Exception as exc:
            body = f"<pre>{h(type(exc).__name__)}: {h(exc)}</pre>".encode("utf-8")
            self.send_response(500)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

    def do_POST(self):
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        if path == "/catalog":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                body = self.rfile.read(length).decode("utf-8", errors="replace")
                params = parse_qs(body, keep_blank_values=True)
                if "catalog_text" in params:
                    catalog_text = params.get("catalog_text", [""])[0].replace("\r\n", "\n")
                    MANUAL_CATALOG.write_text(catalog_text.strip() + "\n", encoding="utf-8")
                else:
                    field_values = {field: params.get(field, []) for field, _ in CATALOG_FORM_FIELDS}
                    max_rows = max((len(values) for values in field_values.values()), default=0)
                    records = []
                    for index in range(max_rows):
                        record = {
                            field: clean_text(values[index] if index < len(values) else "")
                            for field, values in field_values.items()
                        }
                        if any(record.values()):
                            records.append(record)
                    MANUAL_CATALOG.write_text(catalog_records_to_tsv(records), encoding="utf-8")
                self.send_response(303)
                self.send_header("Location", "/catalog?saved=1")
                self.end_headers()
            except Exception as exc:
                body = f"<pre>{h(type(exc).__name__)}: {h(exc)}</pre>".encode("utf-8")
                self.send_response(500)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            return
        if path == "/micobioma":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                body = self.rfile.read(length).decode("utf-8", errors="replace")
                params = parse_qs(body, keep_blank_values=True)
                names = params.get("name", [])
                statuses = params.get("status", [])
                values = params.get("value", [])
                notes = params.get("notes", [])
                records = []
                for index, name in enumerate(names):
                    records.append({
                        "name": name,
                        "status": statuses[index] if index < len(statuses) else "no_evaluado",
                        "value": values[index] if index < len(values) else "",
                        "notes": notes[index] if index < len(notes) else "",
                    })
                save_yeast_profile(records)
                self.send_response(303)
                self.send_header("Location", "/micobioma?saved=1")
                self.end_headers()
            except Exception as exc:
                body = f"<pre>{h(type(exc).__name__)}: {h(exc)}</pre>".encode("utf-8")
                self.send_response(500)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            return
        if path == "/consideraciones":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                body = self.rfile.read(length).decode("utf-8", errors="replace")
                text = parse_qs(body).get("considerations_text", [""])[0].replace("\r\n", "\n")
                CONSIDERATIONS_FILE.write_text(text.strip() + "\n", encoding="utf-8")
                self.send_response(303)
                self.send_header("Location", "/consideraciones?saved=1")
                self.end_headers()
            except Exception as exc:
                body = f"<pre>{h(type(exc).__name__)}: {h(exc)}</pre>".encode("utf-8")
                self.send_response(500)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            return
        if path == "/clinica":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                body = self.rfile.read(length).decode("utf-8", errors="replace")
                params = parse_qs(body, keep_blank_values=True)
                save_clinical_settings({
                    "enterotype_number": params.get("enterotype_number", ["1"])[0],
                    "enterotype_name": params.get("enterotype_name", [""])[0],
                    "enterotype_text": params.get("enterotype_text", [""])[0].replace("\r\n", "\n"),
                    "stool_macro": params.get("stool_macro", [""])[0].replace("\r\n", "\n"),
                    "stool_micro": params.get("stool_micro", [""])[0].replace("\r\n", "\n"),
                })
                save_pathogen_settings({
                    key: params.get(f"pathogen_{key}", ["no_detectado"])[0]
                    for key in DEFAULT_PATHOGEN_SETTINGS
                })
                save_table_header_settings({
                    table_header_key(title): {
                        "patient": params.get(f"table_patient_{table_header_key(title)}", [""])[0],
                        "reference": params.get(f"table_reference_{table_header_key(title)}", [""])[0],
                    }
                    for title, _, _ in TABLE_HEADER_SPECS
                })
                self.send_response(303)
                self.send_header("Location", "/clinica?saved=1")
                self.end_headers()
            except Exception as exc:
                body = f"<pre>{h(type(exc).__name__)}: {h(exc)}</pre>".encode("utf-8")
                self.send_response(500)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            return
        if path != "/upload":
            self.send_error(404)
            return
        try:
            UPLOADS.mkdir(exist_ok=True)
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": self.headers.get("Content-Type", ""),
                    "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                },
            )
            pdf = default_pdf_path()
            xlsx = Path(MANUAL_SOURCE)
            for field_name, fallback in (("pdf_file", default_pdf_path()), ("xlsx_file", Path(MANUAL_SOURCE))):
                item = form[field_name] if field_name in form else None
                if item is not None and getattr(item, "filename", ""):
                    target = UPLOADS / safe_upload_name(item.filename)
                    with target.open("wb") as fh:
                        shutil.copyfileobj(item.file, fh)
                    if field_name == "pdf_file":
                        pdf = target
                    else:
                        xlsx = target
                elif field_name == "pdf_file":
                    pdf = fallback
                else:
                    xlsx = fallback
            save_session_state(pdf=pdf, xlsx=xlsx)
            location = "/?" + urlencode({"pdf": str(pdf), "xlsx": str(xlsx)})
            self.send_response(303)
            self.send_header("Location", location)
            self.end_headers()
        except Exception as exc:
            body = f"<pre>{h(type(exc).__name__)}: {h(exc)}</pre>".encode("utf-8")
            self.send_response(500)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

    def send_file(self, path: Path):
        path = path.resolve()
        if not path.exists() or not path.is_file():
            self.send_error(404)
            return
        suffix = path.suffix.lower()
        content_type = {
            ".css": "text/css; charset=utf-8",
            ".png": "image/png",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
        }.get(suffix, "application/octet-stream")
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


if __name__ == "__main__":
    ensure_data_dir()
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "5055"))
    server = ThreadingHTTPServer((host, port), Handler)
    browser_host = "127.0.0.1" if host in {"0.0.0.0", "::"} else host
    url = f"http://{browser_host}:{port}"
    print(f"Genoma app: {url}")
    if IS_FROZEN and os.environ.get("GENOMA_OPEN_BROWSER", "1") != "0":
        threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    server.serve_forever()
